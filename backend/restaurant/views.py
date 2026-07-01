from datetime import timedelta
from decimal import Decimal
import uuid
import os
import json
import hashlib
import zipfile
import requests
from django.conf import settings
from django.contrib.auth import authenticate, get_user_model
from django.core import signing
from django.core.serializers.json import DjangoJSONEncoder
from django.db import connection
from django.http import FileResponse
from functools import wraps
from django.utils import timezone
from django.db.models import Sum, Count, F, Q
from django.db import transaction
from rest_framework import status, viewsets
from rest_framework.decorators import api_view
from rest_framework.response import Response
from .models import Category, MenuItem, Customer, PhoneVerificationCode, Order, Rider, RestaurantSettings, Coupon, Payment, SmsGatewayMessage, OrderChatMessage, OrderReview, CustomerPushDevice, ExpenseCategory, AccountingSettings, RestaurantFinancialEntry, SystemBackup, Ingredient
from .serializers import CategoryWithItemsSerializer, SendPhoneCodeSerializer, VerifyPhoneCodeSerializer, CustomerSerializer, CreateOrderSerializer, OrderSerializer, RiderSerializer, CategoryAdminSerializer, MenuItemAdminSerializer, MenuItemSerializer, RestaurantSettingsSerializer, CouponSerializer, OrderChatMessageSerializer, OrderReviewSerializer, CustomerPushDeviceSerializer, ExpenseCategorySerializer, AccountingSettingsSerializer, RestaurantFinancialEntrySerializer, SystemBackupSerializer
from .notifications import send_telegram_message, build_order_message, send_customer_order_sms, queue_sms
from .push_notifications import send_order_status_push, send_payment_status_push, send_push_to_phone


ADMIN_TOKEN_SALT = 'casa-de-kebab-admin-v1'
ADMIN_TOKEN_MAX_AGE = 60 * 60 * 12  # 12 hours


def build_admin_token(user):
    return signing.dumps({
        'uid': user.id,
        'username': user.get_username(),
        'is_staff': bool(user.is_staff),
        'is_superuser': bool(user.is_superuser),
    }, salt=ADMIN_TOKEN_SALT)


def get_admin_user_from_request(request):
    auth_header = request.META.get('HTTP_AUTHORIZATION', '') or ''
    token = ''
    if auth_header.lower().startswith('bearer '):
        token = auth_header.split(' ', 1)[1].strip()
    if not token:
        token = request.query_params.get('admin_token', '').strip()
    if not token:
        return None
    try:
        data = signing.loads(token, salt=ADMIN_TOKEN_SALT, max_age=ADMIN_TOKEN_MAX_AGE)
        user = get_user_model().objects.get(id=data.get('uid'), is_active=True)
    except Exception:
        return None
    if not (user.is_staff or user.is_superuser):
        return None
    return user


def admin_token_required(view_func):
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        user = get_admin_user_from_request(request)
        if not user:
            return Response({'detail': 'Admin login required.'}, status=status.HTTP_401_UNAUTHORIZED)
        request.admin_user = user
        return view_func(request, *args, **kwargs)
    return wrapper


@api_view(['POST'])
def admin_login(request):
    username = (request.data.get('username') or '').strip()
    password = request.data.get('password') or ''
    if not username or not password:
        return Response({'detail': 'username and password are required'}, status=status.HTTP_400_BAD_REQUEST)
    user = authenticate(request, username=username, password=password)
    if not user or not user.is_active or not (user.is_staff or user.is_superuser):
        return Response({'detail': 'Invalid admin username or password.'}, status=status.HTTP_403_FORBIDDEN)
    token = build_admin_token(user)
    return Response({
        'success': True,
        'token': token,
        'user': {
            'id': user.id,
            'username': user.get_username(),
            'email': user.email,
            'is_staff': user.is_staff,
            'is_superuser': user.is_superuser,
        },
        'expires_in_hours': 12,
    })


@api_view(['GET'])
def admin_me(request):
    user = get_admin_user_from_request(request)
    if not user:
        return Response({'authenticated': False}, status=status.HTTP_401_UNAUTHORIZED)
    return Response({
        'authenticated': True,
        'user': {
            'id': user.id,
            'username': user.get_username(),
            'email': user.email,
            'is_staff': user.is_staff,
            'is_superuser': user.is_superuser,
        }
    })


RIDER_TOKEN_SALT = 'casa-de-kebab-rider-v1'
RIDER_TOKEN_MAX_AGE = 60 * 60 * 24 * 7


def build_rider_token(rider):
    return signing.dumps({
        'rider_id': rider.id,
        'username': rider.username,
        'phone': rider.phone,
    }, salt=RIDER_TOKEN_SALT)


def get_rider_from_request(request):
    auth_header = request.META.get('HTTP_AUTHORIZATION', '') or ''
    token = ''
    if auth_header.lower().startswith('bearer '):
        token = auth_header.split(' ', 1)[1].strip()
    if not token:
        return None
    try:
        data = signing.loads(token, salt=RIDER_TOKEN_SALT, max_age=RIDER_TOKEN_MAX_AGE)
        return Rider.objects.get(
            id=data.get('rider_id'),
            username=data.get('username'),
            is_active=True,
        )
    except Exception:
        return None


def rider_token_required(view_func):
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        rider = get_rider_from_request(request)
        if not rider:
            return Response(
                {'detail': 'Rider login required or session expired.'},
                status=status.HTTP_401_UNAUTHORIZED,
            )
        request.rider = rider
        return view_func(request, *args, **kwargs)
    return wrapper


@api_view(['POST'])
def rider_login(request):
    username = (request.data.get('username') or '').strip()
    password = request.data.get('password') or ''
    if not username or not password:
        return Response(
            {'detail': 'username and password are required'},
            status=status.HTTP_400_BAD_REQUEST,
        )
    try:
        rider = Rider.objects.get(username=username, is_active=True)
    except Rider.DoesNotExist:
        return Response(
            {'detail': 'Usuario o contraseña incorrectos.'},
            status=status.HTTP_403_FORBIDDEN,
        )
    if not rider.check_password(password):
        return Response(
            {'detail': 'Usuario o contraseña incorrectos.'},
            status=status.HTTP_403_FORBIDDEN,
        )
    return Response({
        'success': True,
        'token': build_rider_token(rider),
        'rider': RiderSerializer(rider).data,
        'expires_in_hours': 168,
    })


@api_view(['GET'])
@rider_token_required
def secure_rider_orders(request):
    rider = request.rider
    qs = rider.orders.exclude(
        status__in=[Order.STATUS_DELIVERED, Order.STATUS_CANCELLED]
    ).prefetch_related('items', 'payments').order_by('-created_at')
    return Response({
        'rider': RiderSerializer(rider).data,
        'orders': OrderSerializer(qs, many=True).data,
    })


@api_view(['POST'])
@rider_token_required
def secure_rider_location(request):
    rider = request.rider
    latitude = request.data.get('latitude')
    longitude = request.data.get('longitude')
    if latitude is None or longitude is None:
        return Response(
            {'detail': 'latitude and longitude are required'},
            status=status.HTTP_400_BAD_REQUEST,
        )
    try:
        lat = float(latitude)
        lng = float(longitude)
    except (TypeError, ValueError):
        return Response({'detail': 'Invalid coordinates'}, status=status.HTTP_400_BAD_REQUEST)

    normal = 40.80 <= lat <= 41.12 and -5.90 <= lng <= -5.35
    swapped = 40.80 <= lng <= 41.12 and -5.90 <= lat <= -5.35
    if swapped:
        lat, lng = lng, lat
        normal = True
    if not normal:
        return Response(
            {'detail': 'La ubicación del repartidor está fuera de Salamanca.'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    rider.current_latitude = round(lat, 7)
    rider.current_longitude = round(lng, 7)
    rider.last_location_at = timezone.now()
    rider.save(update_fields=['current_latitude', 'current_longitude', 'last_location_at'])
    return Response(RiderSerializer(rider).data)


@api_view(['POST'])
@rider_token_required
def secure_rider_update_order_status(request, order_code):
    rider = request.rider
    new_status = (request.data.get('status') or '').strip()
    if not new_status:
        return Response({'detail': 'status is required'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        order = Order.objects.prefetch_related('items', 'payments').get(order_code=order_code)
    except Order.DoesNotExist:
        return Response({'detail': 'Order not found'}, status=status.HTTP_404_NOT_FOUND)

    if order.assigned_rider_id != rider.id:
        return Response(
            {'detail': 'Este pedido no está asignado a este repartidor.'},
            status=status.HTTP_403_FORBIDDEN,
        )

    allowed = {
        Order.STATUS_PENDING: {Order.STATUS_ACCEPTED, Order.STATUS_OUT_FOR_DELIVERY, Order.STATUS_CANCELLED},
        Order.STATUS_ACCEPTED: {Order.STATUS_OUT_FOR_DELIVERY, Order.STATUS_CANCELLED},
        Order.STATUS_PREPARING: {Order.STATUS_OUT_FOR_DELIVERY, Order.STATUS_CANCELLED},
        Order.STATUS_READY: {Order.STATUS_OUT_FOR_DELIVERY, Order.STATUS_CANCELLED},
        Order.STATUS_OUT_FOR_DELIVERY: {Order.STATUS_DELIVERED, Order.STATUS_CANCELLED},
    }

    if new_status not in allowed.get(order.status, set()):
        return Response(
            {'detail': f'Invalid transition: {order.status} -> {new_status}'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    old_status = order.status
    order.status = new_status
    order.save(update_fields=['status', 'updated_at'])
    if old_status != new_status:
        transaction.on_commit(lambda: send_order_status_push(order))
    return Response(OrderSerializer(order).data)


def find_available_rider():
    """Return the active rider with the fewest active delivery orders.

    This keeps assignment simple and predictable for a small restaurant: the
    rider with the smallest number of unfinished delivery orders receives the
    next delivery order. If there are no active riders, return None.
    """
    riders = list(Rider.objects.filter(is_active=True))
    if not riders:
        return None
    active_statuses_to_ignore = [Order.STATUS_DELIVERED, Order.STATUS_CANCELLED]
    return min(
        riders,
        key=lambda rider: rider.orders.filter(delivery_type=Order.DELIVERY_DELIVERY).exclude(status__in=active_statuses_to_ignore).count()
    )


def auto_assign_rider_to_order(order, force_status=False):
    """Assign the best available rider to a delivery order if it has none."""
    if order.delivery_type != Order.DELIVERY_DELIVERY or order.assigned_rider_id:
        return order, False
    rider = find_available_rider()
    if not rider:
        return order, False
    order.assigned_rider = rider
    update_fields = ['assigned_rider', 'updated_at']
    if force_status and order.status in [Order.STATUS_PENDING, Order.STATUS_ACCEPTED, Order.STATUS_PREPARING, Order.STATUS_READY]:
        order.status = Order.STATUS_OUT_FOR_DELIVERY
        update_fields.append('status')
    order.save(update_fields=update_fields)
    return order, True

class MenuViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = CategoryWithItemsSerializer
    def get_queryset(self):
        return Category.objects.filter(is_active=True).prefetch_related('items', 'items__option_groups', 'items__option_groups__options').order_by('sort_order', 'name_es')

@api_view(['POST'])
def send_phone_code(request):
    serializer = SendPhoneCodeSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    phone = serializer.validated_data['phone']
    code = PhoneVerificationCode.generate_code()
    PhoneVerificationCode.objects.filter(phone=phone, is_used=False).update(is_used=True)
    verification = PhoneVerificationCode.objects.create(phone=phone, code=code, expires_at=timezone.now() + timedelta(minutes=5))
    sms_mode = getattr(settings, 'SMS_MODE', 'console')
    if sms_mode == 'console':
        print('======================================')
        print('Casa de Kebab Turco verification code')
        print(f'Phone: {phone}')
        print(f'Code: {verification.code}')
        print('======================================')
    else:
        queue_sms(phone, f'Casa de Kebab Turco: tu código de verificación es {verification.code}. Válido durante 5 minutos.', kind=SmsGatewayMessage.KIND_OTP)
    return Response({'success': True, 'message': 'Verification code sent.', 'mode': sms_mode})

@api_view(['POST'])
def verify_phone_code(request):
    serializer = VerifyPhoneCodeSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    phone = serializer.validated_data['phone']
    code = serializer.validated_data['code']
    try:
        verification = PhoneVerificationCode.objects.filter(phone=phone, code=code, is_used=False).latest('created_at')
    except PhoneVerificationCode.DoesNotExist:
        return Response({'success': False, 'message': 'Invalid code.'}, status=status.HTTP_400_BAD_REQUEST)
    verification.attempt_count += 1
    if verification.is_expired:
        verification.save()
        return Response({'success': False, 'message': 'Code expired.'}, status=status.HTTP_400_BAD_REQUEST)
    verification.is_used = True
    verification.save()
    customer, created = Customer.objects.get_or_create(phone=phone)
    customer.last_login_at = timezone.now()
    customer.save()
    return Response({'success': True, 'customer': CustomerSerializer(customer).data, 'is_new_customer': created})

@api_view(['GET'])
def customer_by_phone(request):
    phone = request.query_params.get('phone')
    if not phone:
        return Response({'detail': 'phone is required'}, status=status.HTTP_400_BAD_REQUEST)
    try:
        customer = Customer.objects.get(phone=phone)
    except Customer.DoesNotExist:
        return Response({'exists': False})
    return Response({'exists': True, 'customer': CustomerSerializer(customer).data})


@api_view(['POST'])
def register_push_device(request):
    phone = ''.join(ch for ch in str(request.data.get('phone') or '') if ch.isdigit())
    token = str(request.data.get('device_token') or '').strip()
    platform = str(request.data.get('platform') or CustomerPushDevice.PLATFORM_ANDROID).strip().lower()
    app_version = str(request.data.get('app_version') or '').strip()[:40]
    customer_id = request.data.get('customer_id')

    if len(phone) < 9 or not token:
        return Response(
            {'detail': 'phone and device_token are required'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    if platform not in dict(CustomerPushDevice.PLATFORM_CHOICES):
        platform = CustomerPushDevice.PLATFORM_ANDROID

    customer = None
    if customer_id:
        customer = Customer.objects.filter(id=customer_id).first()
    if customer is None:
        customer = Customer.objects.filter(phone__endswith=phone[-9:]).first()

    device, created = CustomerPushDevice.objects.update_or_create(
        device_token=token,
        defaults={
            'customer': customer,
            'phone': phone,
            'platform': platform,
            'app_version': app_version,
            'is_active': True,
            'last_seen_at': timezone.now(),
            'last_error': '',
        },
    )
    return Response(
        {
            'success': True,
            'created': created,
            'device': CustomerPushDeviceSerializer(device).data,
        },
        status=status.HTTP_201_CREATED if created else status.HTTP_200_OK,
    )


@api_view(['POST'])
def unregister_push_device(request):
    token = str(request.data.get('device_token') or '').strip()
    phone = ''.join(ch for ch in str(request.data.get('phone') or '') if ch.isdigit())

    if not token:
        return Response(
            {'detail': 'device_token is required'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    queryset = CustomerPushDevice.objects.filter(device_token=token)
    if phone:
        queryset = queryset.filter(phone__endswith=phone[-9:])

    updated = queryset.update(
        is_active=False,
        last_seen_at=timezone.now(),
    )
    return Response({'success': True, 'updated': updated})


@api_view(['POST'])
@admin_token_required
def test_customer_push(request):
    phone = ''.join(ch for ch in str(request.data.get('phone') or '') if ch.isdigit())
    if len(phone) < 9:
        return Response(
            {'detail': 'A valid phone is required'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    result = send_push_to_phone(
        phone,
        'Prueba Casa de Kebab Turco',
        'Las notificaciones push están funcionando correctamente.',
        {'type': 'test'},
    )
    return Response({'success': True, 'result': result})


@api_view(['POST'])
def create_order(request):
    admin_user = get_admin_user_from_request(request)
    serializer = CreateOrderSerializer(
        data=request.data,
        context={
            'request': request,
            'allow_admin_collection': bool(admin_user),
        },
    )
    serializer.is_valid(raise_exception=True)
    order = serializer.save()

    # Auto-assign delivery orders to the currently freest active rider.
    # The order remains pending/accepted until the restaurant changes its status;
    # when it is sent out for delivery, auto assignment is also enforced.
    order, auto_assigned = auto_assign_rider_to_order(order, force_status=False)

    # Send Telegram alert if enabled in .env
    try:
        order = Order.objects.select_related('assigned_rider').prefetch_related('items').get(id=order.id)
        send_telegram_message(build_order_message(order))
    except Exception as exc:
        print(f'Order notification skipped: {exc}')
    sms_sent = False
    try:
        sms_sent = send_customer_order_sms(order)
    except Exception as exc:
        print(f'Customer order SMS skipped: {exc}')

    payload = OrderSerializer(order).data
    payload['auto_assigned_rider'] = auto_assigned
    payload['customer_sms_sent'] = sms_sent
    return Response({'success': True, 'order': payload}, status=status.HTTP_201_CREATED)

@api_view(['GET'])
def order_detail(request, order_code):
    try:
        order = Order.objects.prefetch_related('items', 'payments').get(order_code=order_code)
    except Order.DoesNotExist:
        return Response({'detail': 'Order not found'}, status=status.HTTP_404_NOT_FOUND)
    return Response(OrderSerializer(order).data)


@api_view(['GET'])
def public_order_tracking(request):
    """Allow customers to track an order using order code and phone number."""
    order_code = request.query_params.get('order_code', '').strip().upper()
    phone = request.query_params.get('phone', '').strip()
    if not order_code or not phone:
        return Response({'detail': 'order_code and phone are required'}, status=status.HTTP_400_BAD_REQUEST)
    try:
        order = Order.objects.select_related('assigned_rider').prefetch_related('items', 'payments').get(order_code__iexact=order_code)
    except Order.DoesNotExist:
        return Response({'detail': 'Pedido no encontrado.'}, status=status.HTTP_404_NOT_FOUND)
    clean_input_phone = ''.join(ch for ch in phone if ch.isdigit())
    clean_order_phone = ''.join(ch for ch in order.customer_phone if ch.isdigit())
    if clean_input_phone and clean_order_phone and not clean_order_phone.endswith(clean_input_phone[-9:]):
        return Response({'detail': 'El teléfono no coincide con el pedido.'}, status=status.HTTP_403_FORBIDDEN)
    def normalize_point(latitude, longitude):
        if latitude is None or longitude is None:
            return None, None, False, False
        lat = float(latitude)
        lng = float(longitude)
        normal = 40.80 <= lat <= 41.12 and -5.90 <= lng <= -5.35
        swapped = 40.80 <= lng <= 41.12 and -5.90 <= lat <= -5.35
        if normal:
            return latitude, longitude, True, False
        if swapped:
            return longitude, latitude, True, True
        return None, None, False, False

    order_lat, order_lng, order_valid, order_swapped = normalize_point(
        order.delivery_latitude, order.delivery_longitude
    )
    if order_swapped:
        order.delivery_latitude = order_lat
        order.delivery_longitude = order_lng
        order.save(update_fields=['delivery_latitude', 'delivery_longitude', 'updated_at'])

    rider_corrected = False
    if order.assigned_rider:
        rider_lat, rider_lng, rider_valid, rider_swapped = normalize_point(
            order.assigned_rider.current_latitude,
            order.assigned_rider.current_longitude,
        )
        if rider_swapped:
            order.assigned_rider.current_latitude = rider_lat
            order.assigned_rider.current_longitude = rider_lng
            order.assigned_rider.save(update_fields=['current_latitude', 'current_longitude', 'last_location_at'])
            rider_corrected = True

    payload = OrderSerializer(order).data
    if not order_valid and order.delivery_type == Order.DELIVERY_DELIVERY:
        payload['delivery_latitude'] = None
        payload['delivery_longitude'] = None
        payload['location_warning'] = 'La dirección existe, pero las coordenadas antiguas no son válidas para Salamanca.'

    payload['coordinates_corrected'] = bool(order_swapped)
    payload['rider_coordinates_corrected'] = rider_corrected
    payload['tracking_enabled'] = bool(
        order.assigned_rider
        and order.assigned_rider.current_latitude is not None
        and order.assigned_rider.current_longitude is not None
    )
    payload['restaurant_location'] = {'latitude': '40.974836942683254', 'longitude': '-5.649336331469509'}
    return Response(payload)


@api_view(['GET'])
@admin_token_required
def admin_tracking_orders(request):
    """Return active delivery orders with rider location for the admin live map."""
    qs = Order.objects.select_related('assigned_rider').prefetch_related('items', 'payments').filter(
        delivery_type=Order.DELIVERY_DELIVERY
    ).exclude(status__in=[Order.STATUS_DELIVERED, Order.STATUS_CANCELLED]).order_by('-created_at')[:80]
    return Response(OrderSerializer(qs, many=True).data)


@api_view(['GET'])
def live_orders(request):
    """Return the latest orders for the live restaurant dashboard."""
    limit = int(request.query_params.get('limit', 30))
    qs = Order.objects.prefetch_related('items', 'payments').order_by('-created_at')[:limit]
    return Response(OrderSerializer(qs, many=True).data)


@api_view(['PATCH', 'POST'])
@admin_token_required
def update_order_status(request, order_code):
    try:
        order = Order.objects.get(order_code=order_code)
    except Order.DoesNotExist:
        return Response({'detail': 'Order not found'}, status=status.HTTP_404_NOT_FOUND)

    new_status = request.data.get('status')
    if new_status not in dict(Order.STATUS_CHOICES):
        return Response({'detail': 'Invalid status'}, status=status.HTTP_400_BAD_REQUEST)

    old_status = order.status
    order.status = new_status
    order.save(update_fields=['status', 'updated_at'])
    if old_status != new_status:
        transaction.on_commit(lambda: send_order_status_push(order))
    if new_status == Order.STATUS_DELIVERED:
        consume_inventory_for_order(order)

    # If the restaurant marks a delivery order as ready or out for delivery and
    # no rider is selected yet, automatically choose the freest active rider.
    if new_status in [Order.STATUS_READY, Order.STATUS_OUT_FOR_DELIVERY]:
        order, _ = auto_assign_rider_to_order(order, force_status=(new_status == Order.STATUS_OUT_FOR_DELIVERY))

    return Response(OrderSerializer(order).data)


@api_view(['POST'])
def test_telegram(request):
    ok = send_telegram_message('✅ Test Telegram - Casa de Kebab Turco')
    return Response({'success': ok})


def _parse_boolean(value, field_name='is_active'):
    if isinstance(value, bool):
        return value
    normalized = str(value).strip().lower()
    if normalized in {'true', '1', 'yes', 'on'}:
        return True
    if normalized in {'false', '0', 'no', 'off'}:
        return False
    raise ValueError(f'{field_name} must be true or false')


def _validate_rider_credentials(name, phone, username, password=None, rider=None):
    errors = {}

    if not name:
        errors['name'] = 'El nombre es obligatorio.'
    if not phone:
        errors['phone'] = 'El teléfono es obligatorio.'
    if not username:
        errors['username'] = 'El nombre de usuario es obligatorio.'
    elif len(username) < 3:
        errors['username'] = 'El nombre de usuario debe tener al menos 3 caracteres.'

    if password is not None and password != '' and len(password) < 6:
        errors['password'] = 'La contraseña debe tener al menos 6 caracteres.'

    phone_qs = Rider.objects.filter(phone=phone)
    username_qs = Rider.objects.filter(username=username)
    if rider is not None:
        phone_qs = phone_qs.exclude(id=rider.id)
        username_qs = username_qs.exclude(id=rider.id)

    if phone and phone_qs.exists():
        errors['phone'] = 'Este teléfono ya está registrado para otro repartidor.'
    if username and username_qs.exists():
        errors['username'] = 'Este nombre de usuario ya está en uso.'

    return errors


@api_view(['GET', 'POST'])
@admin_token_required
def riders_list(request):
    """Admin-only list and creation endpoint for rider accounts."""
    if request.method == 'GET':
        qs = Rider.objects.all().order_by('-is_active', 'name')
        return Response(RiderSerializer(qs, many=True).data)

    name = (request.data.get('name') or '').strip()
    phone = (request.data.get('phone') or '').strip()
    username = (request.data.get('username') or '').strip()
    password = request.data.get('password') or ''

    try:
        is_active = _parse_boolean(request.data.get('is_active', True))
    except ValueError as exc:
        return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

    errors = _validate_rider_credentials(
        name=name,
        phone=phone,
        username=username,
        password=password,
    )
    if not password:
        errors['password'] = 'La contraseña es obligatoria para crear el repartidor.'
    if errors:
        return Response(
            {'detail': 'Revisa los datos del repartidor.', 'errors': errors},
            status=status.HTTP_400_BAD_REQUEST,
        )

    rider = Rider(
        name=name,
        phone=phone,
        username=username,
        is_active=is_active,
    )
    rider.set_password(password)
    rider.save()

    return Response(
        RiderSerializer(rider).data,
        status=status.HTTP_201_CREATED,
    )


@api_view(['GET', 'PATCH'])
@admin_token_required
def rider_detail(request, rider_id):
    """Admin-only read/update endpoint, including password and active state."""
    try:
        rider = Rider.objects.get(id=rider_id)
    except Rider.DoesNotExist:
        return Response(
            {'detail': 'Repartidor no encontrado.'},
            status=status.HTTP_404_NOT_FOUND,
        )

    if request.method == 'GET':
        return Response(RiderSerializer(rider).data)

    name = (request.data.get('name', rider.name) or '').strip()
    phone = (request.data.get('phone', rider.phone) or '').strip()
    username = (request.data.get('username', rider.username) or '').strip()
    password = request.data.get('password', '')

    try:
        is_active = _parse_boolean(
            request.data.get('is_active', rider.is_active)
        )
    except ValueError as exc:
        return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

    errors = _validate_rider_credentials(
        name=name,
        phone=phone,
        username=username,
        password=password,
        rider=rider,
    )
    if errors:
        return Response(
            {'detail': 'Revisa los datos del repartidor.', 'errors': errors},
            status=status.HTTP_400_BAD_REQUEST,
        )

    rider.name = name
    rider.phone = phone
    rider.username = username
    rider.is_active = is_active

    update_fields = ['name', 'phone', 'username', 'is_active']
    if password:
        rider.set_password(password)
        update_fields.append('password_hash')

    rider.save(update_fields=update_fields)

    return Response({
        'success': True,
        'message': 'Datos del repartidor actualizados correctamente.',
        'rider': RiderSerializer(rider).data,
    })


@api_view(['POST'])
@admin_token_required
def auto_assign_rider(request, order_code):
    """Manually trigger automatic rider assignment for one delivery order."""
    try:
        order = Order.objects.select_related('assigned_rider').get(order_code=order_code)
    except Order.DoesNotExist:
        return Response({'detail': 'Order not found'}, status=status.HTTP_404_NOT_FOUND)

    if order.delivery_type != Order.DELIVERY_DELIVERY:
        return Response({'detail': 'Auto assignment is only available for delivery orders.'}, status=status.HTTP_400_BAD_REQUEST)

    previous_rider_id = order.assigned_rider_id
    order, assigned = auto_assign_rider_to_order(order, force_status=True)

    # If the order already had a rider, return it as a successful no-op.
    if previous_rider_id and not assigned:
        return Response({
            'success': True,
            'assigned': False,
            'message': 'Order already has a rider.',
            'order': OrderSerializer(order).data,
        })

    if not order.assigned_rider_id:
        return Response({
            'success': False,
            'assigned': False,
            'detail': 'No active rider is available.',
            'order': OrderSerializer(order).data,
        }, status=status.HTTP_409_CONFLICT)

    return Response({
        'success': True,
        'assigned': True,
        'message': 'Rider assigned automatically.',
        'order': OrderSerializer(order).data,
    })


@api_view(['POST'])
@admin_token_required
def assign_rider(request, order_code):
    try:
        order = Order.objects.get(order_code=order_code)
    except Order.DoesNotExist:
        return Response({'detail': 'Order not found'}, status=status.HTTP_404_NOT_FOUND)

    rider_id = request.data.get('rider_id')
    if not rider_id:
        order.assigned_rider = None
        order.save(update_fields=['assigned_rider', 'updated_at'])
        return Response(OrderSerializer(order).data)

    try:
        rider = Rider.objects.get(id=rider_id, is_active=True)
    except Rider.DoesNotExist:
        return Response({'detail': 'Rider not found'}, status=status.HTTP_404_NOT_FOUND)

    order.assigned_rider = rider
    if order.status in [Order.STATUS_PENDING, Order.STATUS_ACCEPTED, Order.STATUS_PREPARING, Order.STATUS_READY]:
        order.status = Order.STATUS_OUT_FOR_DELIVERY
        order.save(update_fields=['assigned_rider', 'status', 'updated_at'])
    else:
        order.save(update_fields=['assigned_rider', 'updated_at'])
    return Response(OrderSerializer(order).data)


@api_view(['GET'])
def rider_orders(request):
    phone = request.query_params.get('phone', '').strip()
    if not phone:
        return Response({'detail': 'phone is required'}, status=status.HTTP_400_BAD_REQUEST)
    try:
        rider = Rider.objects.get(phone=phone, is_active=True)
    except Rider.DoesNotExist:
        return Response({'detail': 'Rider not found'}, status=status.HTTP_404_NOT_FOUND)
    qs = rider.orders.exclude(status__in=[Order.STATUS_DELIVERED, Order.STATUS_CANCELLED]).prefetch_related('items', 'payments').order_by('-created_at')
    return Response({'rider': RiderSerializer(rider).data, 'orders': OrderSerializer(qs, many=True).data})


@api_view(['POST'])
def rider_location(request):
    phone = request.data.get('phone', '').strip()
    latitude = request.data.get('latitude')
    longitude = request.data.get('longitude')
    if not phone or latitude is None or longitude is None:
        return Response({'detail': 'phone, latitude and longitude are required'}, status=status.HTTP_400_BAD_REQUEST)
    try:
        rider = Rider.objects.get(phone=phone, is_active=True)
    except Rider.DoesNotExist:
        return Response({'detail': 'Rider not found'}, status=status.HTTP_404_NOT_FOUND)
    rider.current_latitude = latitude
    rider.current_longitude = longitude
    rider.last_location_at = timezone.now()
    rider.save(update_fields=['current_latitude', 'current_longitude', 'last_location_at'])
    return Response(RiderSerializer(rider).data)


@api_view(['POST'])
def rider_update_order_status(request, order_code):
    phone=(request.data.get('phone') or '').strip(); new_status=(request.data.get('status') or '').strip()
    if not phone or not new_status:return Response({'detail':'phone and status are required'},status=status.HTTP_400_BAD_REQUEST)
    try:rider=Rider.objects.get(phone=phone,is_active=True)
    except Rider.DoesNotExist:return Response({'detail':'Rider not found'},status=status.HTTP_404_NOT_FOUND)
    try:order=Order.objects.select_related('assigned_rider').prefetch_related('items','payments').get(order_code=order_code)
    except Order.DoesNotExist:return Response({'detail':'Order not found'},status=status.HTTP_404_NOT_FOUND)
    if order.assigned_rider_id!=rider.id:return Response({'detail':'Este pedido no está asignado a este repartidor.'},status=status.HTTP_403_FORBIDDEN)
    allowed={Order.STATUS_PENDING:{Order.STATUS_ACCEPTED,Order.STATUS_OUT_FOR_DELIVERY,Order.STATUS_CANCELLED},Order.STATUS_ACCEPTED:{Order.STATUS_OUT_FOR_DELIVERY,Order.STATUS_CANCELLED},Order.STATUS_PREPARING:{Order.STATUS_OUT_FOR_DELIVERY,Order.STATUS_CANCELLED},Order.STATUS_READY:{Order.STATUS_OUT_FOR_DELIVERY,Order.STATUS_CANCELLED},Order.STATUS_OUT_FOR_DELIVERY:{Order.STATUS_DELIVERED,Order.STATUS_CANCELLED}}
    if new_status not in allowed.get(order.status,set()):return Response({'detail':f'No se puede cambiar {order.status} a {new_status}.'},status=status.HTTP_400_BAD_REQUEST)
    old_status=order.status; order.status=new_status; order.save(update_fields=['status','updated_at']);
    if old_status!=new_status: transaction.on_commit(lambda: send_order_status_push(order));return Response(OrderSerializer(order).data)


@api_view(['GET'])
def customer_orders(request):
    """Return a customer profile and all previous orders by phone number."""
    phone = request.query_params.get('phone', '').strip()
    if not phone:
        return Response({'detail': 'phone is required'}, status=status.HTTP_400_BAD_REQUEST)
    try:
        customer = Customer.objects.get(phone=phone)
    except Customer.DoesNotExist:
        return Response({'exists': False, 'orders': []})

    orders = Order.objects.filter(customer=customer).prefetch_related('items', 'payments').order_by('-created_at')
    return Response({
        'exists': True,
        'customer': CustomerSerializer(customer).data,
        'orders': OrderSerializer(orders, many=True).data,
    })


@api_view(['GET'])
@admin_token_required
def dashboard_summary(request):
    """Professional restaurant dashboard: sales, order counts, payments, customers and product performance."""
    today = timezone.localdate()
    today_orders = Order.objects.filter(created_at__date=today)
    all_orders = Order.objects.all()
    paid_orders = all_orders.filter(payment_status=Order.PAYMENT_PAID)

    def safe_total(qs, field='total'):
        return qs.aggregate(value=Sum(field)).get('value') or 0

    payment_breakdown = list(
        all_orders.values('payment_method')
        .annotate(count=Count('id'), total=Sum('total'))
        .order_by('payment_method')
    )

    today_payment_breakdown = list(
        today_orders.values('payment_method')
        .annotate(count=Count('id'), total=Sum('total'))
        .order_by('payment_method')
    )

    status_breakdown = list(
        all_orders.values('status')
        .annotate(count=Count('id'), total=Sum('total'))
        .order_by('status')
    )

    delivery_breakdown = list(
        all_orders.values('delivery_type')
        .annotate(count=Count('id'), total=Sum('total'))
        .order_by('delivery_type')
    )

    product_sales = (
        Order.objects.values('items__name_snapshot')
        .annotate(quantity=Sum('items__quantity'), total=Sum('items__total'))
        .exclude(items__name_snapshot__isnull=True)
        .exclude(items__name_snapshot='')
    )
    top_items = list(product_sales.order_by('-quantity')[:10])
    low_items = list(product_sales.order_by('quantity')[:10])

    today_top_items = list(
        Order.objects.filter(created_at__date=today)
        .values('items__name_snapshot')
        .annotate(quantity=Sum('items__quantity'), total=Sum('items__total'))
        .exclude(items__name_snapshot__isnull=True)
        .exclude(items__name_snapshot='')
        .order_by('-quantity')[:10]
    )

    recent_customers = list(
        Customer.objects.order_by('-last_order_at')[:12]
        .values('id', 'name', 'phone', 'email', 'default_address', 'total_orders', 'last_order_at')
    )

    card_methods = [Order.PAYMENT_CARD_DELIVERY, Order.PAYMENT_ONLINE]
    card_paid = paid_orders.filter(payment_method__in=card_methods)

    return Response({
        'today': str(today),
        'today_sales': safe_total(today_orders),
        'today_orders_count': today_orders.count(),
        'pending_orders_count': all_orders.filter(status=Order.STATUS_PENDING).count(),
        'active_orders_count': all_orders.exclude(status__in=[Order.STATUS_DELIVERED, Order.STATUS_CANCELLED]).count(),
        'delivered_today_count': today_orders.filter(status=Order.STATUS_DELIVERED).count(),
        'cancelled_today_count': today_orders.filter(status=Order.STATUS_CANCELLED).count(),
        'total_sales': safe_total(all_orders),
        'total_orders_count': all_orders.count(),
        'paid_total': safe_total(paid_orders),
        'paid_orders_count': paid_orders.count(),
        'pending_payment_total': safe_total(all_orders.filter(payment_status=Order.PAYMENT_PENDING)),
        'card_paid_total': safe_total(card_paid),
        'card_paid_count': card_paid.count(),
        'cash_total': safe_total(all_orders.filter(payment_method=Order.PAYMENT_CASH)),
        'delivery_fee_total': safe_total(all_orders, 'delivery_fee'),
        'discount_total': safe_total(all_orders, 'discount'),
        'customers_count': Customer.objects.count(),
        'riders_count': Rider.objects.count(),
        'menu_items_count': MenuItem.objects.count(),
        'categories_count': Category.objects.count(),
        'payment_breakdown': payment_breakdown,
        'today_payment_breakdown': today_payment_breakdown,
        'status_breakdown': status_breakdown,
        'delivery_breakdown': delivery_breakdown,
        'top_items': top_items,
        'today_top_items': today_top_items,
        'low_items': low_items,
        'recent_customers': recent_customers,
    })



@api_view(['GET'])
@admin_token_required
def admin_customers(request):
    """Return customers for the professional admin panel."""
    limit = int(request.GET.get('limit', 200))
    qs = Customer.objects.order_by('-last_order_at', '-created_at')[:limit]
    rows = []
    for c in qs:
        orders = c.orders.all()
        rows.append({
            'id': c.id,
            'name': c.name,
            'phone': c.phone,
            'email': c.email,
            'default_address': c.default_address,
            'total_orders': c.total_orders,
            'last_order_at': c.last_order_at,
            'created_at': c.created_at,
            'total_spent': orders.aggregate(value=Sum('total')).get('value') or 0,
            'last_status': orders.order_by('-created_at').values_list('status', flat=True).first() or '',
        })
    return Response(rows)


@api_view(['PATCH', 'POST'])
@admin_token_required
def update_payment_status(request, order_code):
    try:
        order = Order.objects.get(order_code=order_code)
    except Order.DoesNotExist:
        return Response({'detail': 'Order not found'}, status=status.HTTP_404_NOT_FOUND)

    new_status = request.data.get('payment_status')
    if new_status not in dict(Order.PAYMENT_STATUS_CHOICES):
        return Response({'detail': 'Invalid payment status'}, status=status.HTTP_400_BAD_REQUEST)

    old_payment_status = order.payment_status
    order.payment_status = new_status
    order.save(update_fields=['payment_status', 'updated_at'])
    order.payments.update(status=new_status)
    if old_payment_status != new_status:
        transaction.on_commit(lambda: send_payment_status_push(order))
    return Response(OrderSerializer(order).data)


# =========================
# Menu Management API
# =========================

@api_view(['GET', 'POST'])
@admin_token_required
def admin_categories(request):
    if request.method == 'GET':
        qs = Category.objects.all().order_by('sort_order', 'name_es')
        return Response(CategoryAdminSerializer(qs, many=True).data)

    serializer = CategoryAdminSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    category = serializer.save()
    return Response(CategoryAdminSerializer(category).data, status=status.HTTP_201_CREATED)


@api_view(['PATCH', 'DELETE'])
@admin_token_required
def admin_category_detail(request, category_id):
    try:
        category = Category.objects.get(id=category_id)
    except Category.DoesNotExist:
        return Response({'detail': 'Category not found'}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'DELETE':
        category.is_active = False
        category.save(update_fields=['is_active'])
        return Response({'success': True, 'message': 'Category disabled'})

    serializer = CategoryAdminSerializer(category, data=request.data, partial=True)
    serializer.is_valid(raise_exception=True)
    category = serializer.save()
    return Response(CategoryAdminSerializer(category).data)


@api_view(['GET', 'POST'])
@admin_token_required
def admin_menu_items(request):
    if request.method == 'GET':
        qs = MenuItem.objects.select_related('category').all().order_by('category__sort_order', 'sort_order', 'name_es')
        return Response(MenuItemAdminSerializer(qs, many=True, context={'request': request}).data)

    serializer = MenuItemAdminSerializer(data=request.data, context={'request': request})
    serializer.is_valid(raise_exception=True)
    item = serializer.save()
    return Response(MenuItemAdminSerializer(item, context={'request': request}).data, status=status.HTTP_201_CREATED)


@api_view(['PATCH', 'DELETE'])
@admin_token_required
def admin_menu_item_detail(request, item_id):
    try:
        item = MenuItem.objects.get(id=item_id)
    except MenuItem.DoesNotExist:
        return Response({'detail': 'Menu item not found'}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'DELETE':
        item.is_active = False
        item.is_available = False
        item.save(update_fields=['is_active', 'is_available', 'updated_at'])
        return Response({'success': True, 'message': 'Menu item archived'})

    serializer = MenuItemAdminSerializer(item, data=request.data, partial=True, context={'request': request})
    serializer.is_valid(raise_exception=True)
    item = serializer.save()
    return Response(MenuItemAdminSerializer(item, context={'request': request}).data)


@api_view(['POST'])
@admin_token_required
def admin_menu_item_image(request, item_id):
    try:
        item = MenuItem.objects.get(id=item_id)
    except MenuItem.DoesNotExist:
        return Response({'detail': 'Menu item not found'}, status=status.HTTP_404_NOT_FOUND)

    image = request.FILES.get('image')
    if not image:
        return Response({'detail': 'image file is required'}, status=status.HTTP_400_BAD_REQUEST)

    item.image = image
    item.save(update_fields=['image', 'updated_at'])
    return Response(MenuItemAdminSerializer(item, context={'request': request}).data)




@api_view(['GET'])
def google_place_details(request):
    place_id = (request.query_params.get('place_id') or '').strip()
    if not place_id:
        return Response({'detail': 'place_id is required'}, status=status.HTTP_400_BAD_REQUEST)

    api_key = getattr(settings, 'GOOGLE_PLACES_API_KEY', '')
    if not api_key:
        return Response({'detail': 'GOOGLE_PLACES_API_KEY is not configured'}, status=status.HTTP_503_SERVICE_UNAVAILABLE)

    url = f'https://places.googleapis.com/v1/places/{place_id}'
    headers = {
        'X-Goog-Api-Key': api_key,
        'X-Goog-FieldMask': 'id,formattedAddress,location',
    }

    try:
        response = requests.get(url, headers=headers, timeout=8)
        if response.status_code >= 400:
            return Response({'detail': response.text[:300]}, status=status.HTTP_400_BAD_REQUEST)

        data = response.json()
        location = data.get('location') or {}
        latitude = location.get('latitude')
        longitude = location.get('longitude')
        formatted_address = data.get('formattedAddress') or ''

        if latitude is None or longitude is None:
            return Response({'detail': 'Google Places did not return coordinates.'}, status=status.HTTP_400_BAD_REQUEST)

        latitude = float(latitude)
        longitude = float(longitude)

        if not (40.80 <= latitude <= 41.12 and -5.90 <= longitude <= -5.35):
            return Response({'detail': 'La dirección seleccionada está fuera de Salamanca.'}, status=status.HTTP_400_BAD_REQUEST)

        return Response({
            'place_id': place_id,
            'formatted_address': formatted_address.replace(', Spain', ', España'),
            'latitude': latitude,
            'longitude': longitude,
        })
    except Exception as exc:
        return Response({'detail': str(exc)}, status=status.HTTP_502_BAD_GATEWAY)


@api_view(['GET'])
def google_places_autocomplete(request):
    """Google Places autocomplete enriched with coordinates for the customer app."""
    q = (request.query_params.get('q') or '').strip()
    if len(q) < 1:
        return Response({'predictions': []})

    api_key = getattr(settings, 'GOOGLE_PLACES_API_KEY', '')
    if not api_key:
        return Response(
            {'predictions': [], 'detail': 'GOOGLE_PLACES_API_KEY is not configured on backend.'},
            status=status.HTTP_200_OK
        )

    autocomplete_url = 'https://places.googleapis.com/v1/places:autocomplete'
    autocomplete_payload = {
        'input': q,
        'languageCode': 'es',
        'regionCode': 'ES',
        'includedRegionCodes': ['ES'],
        'locationBias': {
            'circle': {
                'center': {
                    'latitude': 40.974836942683254,
                    'longitude': -5.649336331469509
                },
                'radius': 12000.0,
            }
        },
    }
    autocomplete_headers = {
        'Content-Type': 'application/json',
        'X-Goog-Api-Key': api_key,
        'X-Goog-FieldMask': (
            'suggestions.placePrediction.placeId,'
            'suggestions.placePrediction.text,'
            'suggestions.placePrediction.structuredFormat'
        ),
    }

    try:
        response = requests.post(
            autocomplete_url,
            json=autocomplete_payload,
            headers=autocomplete_headers,
            timeout=8
        )
        if response.status_code >= 400:
            return Response(
                {'predictions': [], 'detail': response.text[:300]},
                status=status.HTTP_200_OK
            )
        data = response.json()
    except Exception as exc:
        return Response(
            {'predictions': [], 'detail': str(exc)},
            status=status.HTTP_200_OK
        )

    predictions = []

    for item in data.get('suggestions', [])[:8]:
        prediction = item.get('placePrediction') or {}
        place_id = prediction.get('placeId') or ''
        text_obj = prediction.get('text') or {}
        structured = prediction.get('structuredFormat') or {}

        main_text = (
            (structured.get('mainText') or {}).get('text')
            or text_obj.get('text')
            or ''
        )
        secondary_text = (
            (structured.get('secondaryText') or {}).get('text')
            or ''
        )
        description = (
            text_obj.get('text')
            or ', '.join(x for x in [main_text, secondary_text] if x)
        )

        if not description or not place_id:
            continue

        latitude = None
        longitude = None
        formatted_address = description

        try:
            details_url = f'https://places.googleapis.com/v1/places/{place_id}'
            details_headers = {
                'X-Goog-Api-Key': api_key,
                'X-Goog-FieldMask': 'id,formattedAddress,location',
            }
            details_response = requests.get(
                details_url,
                headers=details_headers,
                timeout=8
            )

            if details_response.status_code < 400:
                details = details_response.json()
                location = details.get('location') or {}
                latitude = location.get('latitude')
                longitude = location.get('longitude')
                formatted_address = (
                    details.get('formattedAddress')
                    or description
                )
        except Exception:
            pass

        if latitude is None or longitude is None:
            continue

        latitude = float(latitude)
        longitude = float(longitude)

        if not (
            40.80 <= latitude <= 41.12
            and -5.90 <= longitude <= -5.35
        ):
            continue

        predictions.append({
            'place_id': place_id,
            'description': formatted_address.replace(', Spain', ', España'),
            'main_text': main_text,
            'secondary_text': secondary_text.replace(', Spain', ', España'),
            'latitude': latitude,
            'longitude': longitude,
        })

    return Response({'predictions': predictions[:8]})

@api_view(['GET'])
def public_settings(request):
    settings_obj = RestaurantSettings.current()
    return Response(RestaurantSettingsSerializer(settings_obj).data)


@api_view(['GET', 'PATCH'])
@admin_token_required
def admin_restaurant_settings(request):
    settings_obj = RestaurantSettings.current()
    if request.method == 'GET':
        return Response(RestaurantSettingsSerializer(settings_obj).data)

    serializer = RestaurantSettingsSerializer(settings_obj, data=request.data, partial=True)
    serializer.is_valid(raise_exception=True)
    serializer.save()
    return Response(serializer.data)


@api_view(['POST'])
def validate_coupon(request):
    code = request.data.get('code', '').strip().upper()
    subtotal = Decimal(str(request.data.get('subtotal', '0.00') or '0.00'))
    phone = request.data.get('phone', '').strip()
    customer = None
    if phone:
        customer = Customer.objects.filter(phone=phone).first()
    if not code:
        return Response({'valid': False, 'message': 'Introduce un código.'}, status=status.HTTP_400_BAD_REQUEST)
    try:
        coupon = Coupon.objects.get(code__iexact=code)
    except Coupon.DoesNotExist:
        return Response({'valid': False, 'message': 'Cupón no encontrado.'}, status=status.HTTP_404_NOT_FOUND)
    valid, message = coupon.is_valid_for(subtotal, customer=customer)
    discount = coupon.calculate_discount(subtotal) if valid else Decimal('0.00')
    return Response({
        'valid': valid,
        'message': message,
        'code': coupon.code,
        'discount': str(discount),
        'coupon': CouponSerializer(coupon).data,
    })


@api_view(['GET', 'POST'])
@admin_token_required
def admin_coupons(request):
    if request.method == 'GET':
        qs = Coupon.objects.all().order_by('code')
        return Response(CouponSerializer(qs, many=True).data)
    serializer = CouponSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    serializer.save(code=serializer.validated_data['code'].upper())
    return Response(serializer.data, status=status.HTTP_201_CREATED)


@api_view(['PATCH', 'DELETE'])
@admin_token_required
def admin_coupon_detail(request, coupon_id):
    try:
        coupon = Coupon.objects.get(id=coupon_id)
    except Coupon.DoesNotExist:
        return Response({'detail': 'Coupon not found'}, status=status.HTTP_404_NOT_FOUND)
    if request.method == 'DELETE':
        coupon.is_active = False
        coupon.save(update_fields=['is_active'])
        return Response({'success': True})
    data = request.data.copy()
    if 'code' in data:
        data['code'] = str(data['code']).upper()
    serializer = CouponSerializer(coupon, data=data, partial=True)
    serializer.is_valid(raise_exception=True)
    serializer.save()
    return Response(serializer.data)


# =========================
# Demo Online Payment API
# =========================

@api_view(['POST'])
def create_online_payment(request, order_code):
    """Create a demo online payment session.

    This is a safe test payment flow. It does not charge real money.
    Later we can replace provider='demo' with Stripe or Redsys/BBVA.
    """
    try:
        order = Order.objects.get(order_code=order_code)
    except Order.DoesNotExist:
        return Response({'detail': 'Order not found'}, status=status.HTTP_404_NOT_FOUND)

    if order.payment_method != Order.PAYMENT_ONLINE:
        order.payment_method = Order.PAYMENT_ONLINE
        order.payment_status = Order.PAYMENT_PENDING
        order.save(update_fields=['payment_method', 'payment_status', 'updated_at'])

    payment = order.payments.filter(provider='demo').order_by('-created_at').first()
    if not payment:
        payment = Payment.objects.create(
            order=order,
            method=Order.PAYMENT_ONLINE,
            provider='demo',
            amount=order.total,
            status=Order.PAYMENT_PENDING,
            transaction_id=f'DEMO-{uuid.uuid4().hex[:12].upper()}',
            raw_response={'mode': 'demo', 'created_by': 'Casa de Kebab Turco local test'},
        )

    return Response({
        'success': True,
        'provider': 'demo',
        'order_code': order.order_code,
        'amount': str(order.total),
        'payment_status': order.payment_status,
        'transaction_id': payment.transaction_id,
        'checkout_url': f'/payment-demo/{order.order_code}',
        'message': 'Demo payment session created. No real money is charged.',
    })


@api_view(['POST'])
def confirm_online_payment(request, order_code):
    """Confirm or fail a demo online payment.

    Body: {"result": "success"} or {"result": "failed"}
    """
    result = request.data.get('result', 'success')

    try:
        order = Order.objects.get(order_code=order_code)
    except Order.DoesNotExist:
        return Response({'detail': 'Order not found'}, status=status.HTTP_404_NOT_FOUND)

    new_status = Order.PAYMENT_PAID if result == 'success' else Order.PAYMENT_FAILED
    order.payment_method = Order.PAYMENT_ONLINE
    order.payment_status = new_status
    order.save(update_fields=['payment_method', 'payment_status', 'updated_at'])

    payment = order.payments.filter(provider='demo').order_by('-created_at').first()
    if not payment:
        payment = Payment.objects.create(
            order=order,
            method=Order.PAYMENT_ONLINE,
            provider='demo',
            amount=order.total,
            transaction_id=f'DEMO-{uuid.uuid4().hex[:12].upper()}',
        )
    payment.status = new_status
    payment.raw_response = {'mode': 'demo', 'result': result, 'confirmed_at': timezone.now().isoformat()}
    payment.save(update_fields=['status', 'raw_response'])

    return Response({
        'success': True,
        'order': OrderSerializer(order).data,
        'message': 'Payment updated successfully.' if result == 'success' else 'Payment marked as failed.',
    })


@api_view(['GET'])
def online_payment_status(request, order_code):
    try:
        order = Order.objects.prefetch_related('items', 'payments').get(order_code=order_code)
    except Order.DoesNotExist:
        return Response({'detail': 'Order not found'}, status=status.HTTP_404_NOT_FOUND)
    return Response(OrderSerializer(order).data)




def _digits(value):
    return ''.join(ch for ch in str(value or '') if ch.isdigit())


def _phone_matches(left, right):
    a, b = _digits(left), _digits(right)
    return bool(a and b and a[-9:] == b[-9:])


def _get_customer_order(order_code, phone):
    try:
        order = Order.objects.select_related('assigned_rider').get(order_code__iexact=str(order_code or '').strip())
    except Order.DoesNotExist:
        return None, Response({'detail': 'Pedido no encontrado.'}, status=status.HTTP_404_NOT_FOUND)
    if not _phone_matches(order.customer_phone, phone):
        return None, Response({'detail': 'El teléfono no coincide con el pedido.'}, status=status.HTTP_403_FORBIDDEN)
    return order, None


@api_view(['GET'])
def order_tracking_location(request, order_code):
    order, error = _get_customer_order(order_code, request.query_params.get('phone'))
    if error:
        return error
    rider = order.assigned_rider
    location = None
    if order.status == Order.STATUS_OUT_FOR_DELIVERY and rider and rider.current_latitude is not None and rider.current_longitude is not None:
        location = {
            'latitude': str(rider.current_latitude),
            'longitude': str(rider.current_longitude),
            'updated_at': rider.last_location_at,
            'rider_name': rider.name,
        }
    return Response({
        'order_code': order.order_code,
        'status': order.status,
        'location': location,
        'delivered': order.status == Order.STATUS_DELIVERED,
    })


def _chat_actor_allowed(request, order, sender_type):
    if sender_type == OrderChatMessage.SENDER_ADMIN:
        user = get_admin_user_from_request(request)
        return bool(user), (user.get_username() if user else '')
    actor_phone = request.data.get('phone') if request.method == 'POST' else request.query_params.get('phone')
    if sender_type == OrderChatMessage.SENDER_CUSTOMER:
        return _phone_matches(order.customer_phone, actor_phone), order.customer_name or 'Cliente'
    if sender_type == OrderChatMessage.SENDER_RIDER:
        rider = order.assigned_rider
        return bool(rider and _phone_matches(rider.phone, actor_phone)), (rider.name if rider else '')
    return False, ''


@api_view(['GET', 'POST'])
def order_chat(request, order_code):
    try:
        order = Order.objects.select_related('assigned_rider').get(order_code__iexact=order_code)
    except Order.DoesNotExist:
        return Response({'detail': 'Pedido no encontrado.'}, status=status.HTTP_404_NOT_FOUND)

    sender_type = str((request.data if request.method == 'POST' else request.query_params).get('sender_type', 'customer')).strip().lower()
    allowed, sender_name = _chat_actor_allowed(request, order, sender_type)
    if not allowed:
        return Response({'detail': 'No autorizado para este chat.'}, status=status.HTTP_403_FORBIDDEN)

    is_admin = sender_type == OrderChatMessage.SENDER_ADMIN
    if order.status == Order.STATUS_DELIVERED and not is_admin:
        return Response({'messages': [], 'chat_closed': True, 'detail': 'El chat se oculta después de la entrega.'})

    if request.method == 'GET':
        rows = order.chat_messages.all()
        return Response({'messages': OrderChatMessageSerializer(rows, many=True).data, 'chat_closed': False})

    message = str(request.data.get('message', '') or '').strip()
    if not message:
        return Response({'detail': 'El mensaje está vacío.'}, status=status.HTTP_400_BAD_REQUEST)
    row = OrderChatMessage.objects.create(order=order, sender_type=sender_type, sender_name=sender_name, message=message[:1200])
    return Response(OrderChatMessageSerializer(row).data, status=status.HTTP_201_CREATED)


@api_view(['POST'])
def create_order_review(request):
    order, error = _get_customer_order(request.data.get('order_code'), request.data.get('phone'))
    if error:
        return error
    if order.status != Order.STATUS_DELIVERED:
        return Response({'detail': 'La opinión se puede enviar después de la entrega.'}, status=status.HTTP_400_BAD_REQUEST)
    if hasattr(order, 'review'):
        return Response({'detail': 'Ya existe una opinión para este pedido.'}, status=status.HTTP_409_CONFLICT)
    try:
        rating = int(request.data.get('rating', 0))
    except (TypeError, ValueError):
        rating = 0
    comment = str(request.data.get('comment', '') or '').strip()
    if rating < 1 or rating > 5 or not comment:
        return Response({'detail': 'Selecciona de 1 a 5 estrellas y escribe tu opinión.'}, status=status.HTTP_400_BAD_REQUEST)
    row = OrderReview.objects.create(
        order=order,
        customer_name=order.customer_name or 'Cliente',
        customer_phone=order.customer_phone,
        rating=rating,
        comment=comment[:1200],
    )
    return Response({'success': True, 'message': 'Opinión enviada y pendiente de aprobación.', 'review': OrderReviewSerializer(row).data}, status=status.HTTP_201_CREATED)


@api_view(['GET'])
def public_reviews(request):
    rows = OrderReview.objects.filter(status=OrderReview.STATUS_APPROVED).select_related('order').order_by('-approved_at', '-created_at')[:30]
    return Response(OrderReviewSerializer(rows, many=True).data)

def _gateway_authorized(request):
    expected = str(getattr(settings, 'SMS_GATEWAY_TOKEN', '') or '').strip()
    auth = str(request.META.get('HTTP_AUTHORIZATION', '') or '')
    provided = auth.split(' ', 1)[1].strip() if auth.lower().startswith('bearer ') else ''
    return bool(expected and provided and provided == expected)


@api_view(['GET'])
def sms_gateway_pending(request):
    if not _gateway_authorized(request):
        return Response({'detail': 'Invalid gateway token.'}, status=status.HTTP_401_UNAUTHORIZED)
    device_id = str(request.query_params.get('device_id', '') or '').strip()
    limit = min(max(int(request.query_params.get('limit', 20)), 1), 50)
    with transaction.atomic():
        rows = list(
            SmsGatewayMessage.objects.select_for_update(skip_locked=True)
            .filter(status=SmsGatewayMessage.STATUS_PENDING)
            .order_by('created_at')[:limit]
        )
        ids = [row.id for row in rows]
        if ids:
            SmsGatewayMessage.objects.filter(id__in=ids).update(
                status=SmsGatewayMessage.STATUS_PROCESSING,
                device_id=device_id,
                attempts=F('attempts') + 1,
            )
    return Response({'messages': [
        {'id': row.id, 'phone': row.phone, 'message': row.message}
        for row in rows
    ]})


@api_view(['POST'])
def sms_gateway_mark(request):
    if not _gateway_authorized(request):
        return Response({'detail': 'Invalid gateway token.'}, status=status.HTTP_401_UNAUTHORIZED)
    message_id = request.data.get('id')
    new_status = str(request.data.get('status', '') or '').strip().lower()
    if new_status not in [SmsGatewayMessage.STATUS_SENT, SmsGatewayMessage.STATUS_FAILED]:
        return Response({'detail': 'Invalid status.'}, status=status.HTTP_400_BAD_REQUEST)
    try:
        row = SmsGatewayMessage.objects.get(id=message_id)
    except SmsGatewayMessage.DoesNotExist:
        return Response({'detail': 'Message not found.'}, status=status.HTTP_404_NOT_FOUND)
    row.status = new_status
    row.error = str(request.data.get('error', '') or '')
    row.device_id = str(request.data.get('device_id', '') or row.device_id)
    if new_status == SmsGatewayMessage.STATUS_SENT:
        row.sent_at = timezone.now()
    row.save(update_fields=['status', 'error', 'device_id', 'sent_at', 'updated_at'])
    return Response({'success': True, 'id': row.id, 'status': row.status})

# =========================
# Partner accounting API
# =========================

def _money_sum(queryset, field='amount'):
    return queryset.aggregate(value=Sum(field)).get('value') or Decimal('0.00')


def _accounting_summary_payload():
    settings_obj = AccountingSettings.current()
    approved = RestaurantFinancialEntry.objects.filter(
        status__in=[
            RestaurantFinancialEntry.STATUS_APPROVED,
            RestaurantFinancialEntry.STATUS_REIMBURSED,
        ]
    )
    expenses = approved.filter(entry_type=RestaurantFinancialEntry.TYPE_EXPENSE)
    contributions = approved.filter(entry_type=RestaurantFinancialEntry.TYPE_CONTRIBUTION)
    settlements = approved.filter(entry_type=RestaurantFinancialEntry.TYPE_SETTLEMENT)

    saeid_expenses = _money_sum(expenses.filter(paid_by=RestaurantFinancialEntry.PARTY_SAEID))
    ahmed_expenses = _money_sum(expenses.filter(paid_by=RestaurantFinancialEntry.PARTY_AHMED))
    bbva_expenses = _money_sum(expenses.filter(paid_by=RestaurantFinancialEntry.PARTY_BBVA))

    saeid_contributions = _money_sum(
        contributions.filter(contribution_from=RestaurantFinancialEntry.PARTY_SAEID)
    )
    ahmed_contributions = _money_sum(
        contributions.filter(contribution_from=RestaurantFinancialEntry.PARTY_AHMED)
    )

    settlements_saeid_to_ahmed = _money_sum(
        settlements.filter(
            paid_by=RestaurantFinancialEntry.PARTY_SAEID,
            settlement_to=RestaurantFinancialEntry.PARTY_AHMED,
        )
    )
    settlements_ahmed_to_saeid = _money_sum(
        settlements.filter(
            paid_by=RestaurantFinancialEntry.PARTY_AHMED,
            settlement_to=RestaurantFinancialEntry.PARTY_SAEID,
        )
    )

    personal_total = saeid_expenses + ahmed_expenses
    saeid_target = (
        personal_total * settings_obj.saeid_share_percent / Decimal('100.00')
    )
    ahmed_target = (
        personal_total * settings_obj.ahmed_share_percent / Decimal('100.00')
    )

    # Positive means Saeid should receive money; negative means Ahmed should receive.
    raw_saeid_credit = saeid_expenses - saeid_target
    settlement_net_to_saeid = settlements_ahmed_to_saeid - settlements_saeid_to_ahmed
    saeid_credit_after_settlement = raw_saeid_credit - settlement_net_to_saeid

    if saeid_credit_after_settlement > 0:
        settlement = {
            'debtor': 'Ahmed',
            'creditor': 'Saeid',
            'amount': str(saeid_credit_after_settlement.quantize(Decimal('0.01'))),
        }
    elif saeid_credit_after_settlement < 0:
        settlement = {
            'debtor': 'Saeid',
            'creditor': 'Ahmed',
            'amount': str(abs(saeid_credit_after_settlement).quantize(Decimal('0.01'))),
        }
    else:
        settlement = {'debtor': '', 'creditor': '', 'amount': '0.00'}

    bbva_balance = (
        settings_obj.bbva_initial_balance
        + saeid_contributions
        + ahmed_contributions
        - bbva_expenses
    )

    month_start = timezone.localdate().replace(day=1)
    month_expenses = _money_sum(expenses.filter(entry_date__gte=month_start))

    by_category = list(
        expenses.values('category__name')
        .annotate(total=Sum('amount'), count=Count('id'))
        .order_by('-total')
    )
    for row in by_category:
        row['name'] = row.pop('category__name') or 'Sin categoría'
        row['total'] = str(row['total'] or Decimal('0.00'))

    return {
        'settings': AccountingSettingsSerializer(settings_obj).data,
        'total_expenses': str(_money_sum(expenses)),
        'month_expenses': str(month_expenses),
        'saeid_expenses': str(saeid_expenses),
        'ahmed_expenses': str(ahmed_expenses),
        'bbva_expenses': str(bbva_expenses),
        'saeid_contributions': str(saeid_contributions),
        'ahmed_contributions': str(ahmed_contributions),
        'bbva_balance': str(bbva_balance),
        'personal_expenses_total': str(personal_total),
        'saeid_target_share': str(saeid_target.quantize(Decimal('0.01'))),
        'ahmed_target_share': str(ahmed_target.quantize(Decimal('0.01'))),
        'settlement': settlement,
        'by_category': by_category,
    }


@api_view(['GET'])
@admin_token_required
def admin_accounting_summary(request):
    return Response(_accounting_summary_payload())


@api_view(['GET', 'PATCH'])
@admin_token_required
def admin_accounting_settings(request):
    settings_obj = AccountingSettings.current()
    if request.method == 'GET':
        return Response(AccountingSettingsSerializer(settings_obj).data)

    serializer = AccountingSettingsSerializer(
        settings_obj,
        data=request.data,
        partial=True,
    )
    serializer.is_valid(raise_exception=True)
    serializer.save()
    return Response(serializer.data)


@api_view(['GET', 'POST'])
@admin_token_required
def admin_expense_categories(request):
    if request.method == 'GET':
        qs = ExpenseCategory.objects.all().order_by('sort_order', 'name')
        return Response(ExpenseCategorySerializer(qs, many=True).data)

    serializer = ExpenseCategorySerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    category = serializer.save()
    return Response(
        ExpenseCategorySerializer(category).data,
        status=status.HTTP_201_CREATED,
    )


@api_view(['GET', 'POST'])
@admin_token_required
def admin_financial_entries(request):
    if request.method == 'GET':
        qs = RestaurantFinancialEntry.objects.select_related('category').all()

        entry_type = (request.query_params.get('entry_type') or '').strip()
        paid_by = (request.query_params.get('paid_by') or '').strip()
        date_from = (request.query_params.get('date_from') or '').strip()
        date_to = (request.query_params.get('date_to') or '').strip()
        search = (request.query_params.get('search') or '').strip()

        if entry_type:
            qs = qs.filter(entry_type=entry_type)
        if paid_by:
            qs = qs.filter(paid_by=paid_by)
        if date_from:
            qs = qs.filter(entry_date__gte=date_from)
        if date_to:
            qs = qs.filter(entry_date__lte=date_to)
        if search:
            qs = qs.filter(
                Q(title__icontains=search)
                | Q(description__icontains=search)
                | Q(invoice_number__icontains=search)
                | Q(bank_reference__icontains=search)
            )

        limit = min(int(request.query_params.get('limit', 500)), 1000)
        serializer = RestaurantFinancialEntrySerializer(
            qs[:limit],
            many=True,
            context={'request': request},
        )
        return Response(serializer.data)

    serializer = RestaurantFinancialEntrySerializer(
        data=request.data,
        context={'request': request},
    )
    serializer.is_valid(raise_exception=True)
    entry = serializer.save(
        created_by_username=request.admin_user.get_username(),
        updated_by_username=request.admin_user.get_username(),
    )
    return Response(
        RestaurantFinancialEntrySerializer(
            entry,
            context={'request': request},
        ).data,
        status=status.HTTP_201_CREATED,
    )


@api_view(['GET', 'PATCH', 'DELETE'])
@admin_token_required
def admin_financial_entry_detail(request, entry_id):
    try:
        entry = RestaurantFinancialEntry.objects.select_related('category').get(id=entry_id)
    except RestaurantFinancialEntry.DoesNotExist:
        return Response(
            {'detail': 'Movimiento financiero no encontrado.'},
            status=status.HTTP_404_NOT_FOUND,
        )

    if request.method == 'GET':
        return Response(
            RestaurantFinancialEntrySerializer(
                entry,
                context={'request': request},
            ).data
        )

    if request.method == 'DELETE':
        entry.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    serializer = RestaurantFinancialEntrySerializer(
        entry,
        data=request.data,
        partial=True,
        context={'request': request},
    )
    serializer.is_valid(raise_exception=True)
    entry = serializer.save(
        updated_by_username=request.admin_user.get_username(),
    )
    return Response(
        RestaurantFinancialEntrySerializer(
            entry,
            context={'request': request},
        ).data
    )

# =========================
# System backup and health
# =========================

BACKUP_DIRECTORY_NAME = 'system_backups'


def _backup_root():
    root = os.path.join(str(settings.MEDIA_ROOT), BACKUP_DIRECTORY_NAME)
    os.makedirs(root, exist_ok=True)
    return root


def _sha256_file(file_path):
    digest = hashlib.sha256()
    with open(file_path, 'rb') as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b''):
            digest.update(chunk)
    return digest.hexdigest()


def _json_safe_rows(queryset, fields):
    rows = []
    for row in queryset.values(*fields):
        rows.append(row)
    return rows


def _database_backup_payload():
    """Create a portable, password-free JSON export of business data."""
    from .models import (
        MenuOptionGroup, MenuOption, CustomerAddress, OrderItem,
        ExpenseCategory, AccountingSettings, RestaurantFinancialEntry,
    )

    return {
        'metadata': {
            'project': 'Casa de Kebab Turco',
            'created_at': timezone.now(),
            'format_version': 1,
            'warning': (
                'This is a safe business-data JSON export. '
                'Passwords, Django users, API tokens and secrets are excluded.'
            ),
        },
        'categories': _json_safe_rows(
            Category.objects.all(),
            ['id', 'name_es', 'name_en', 'slug', 'sort_order', 'is_active'],
        ),
        'menu_items': _json_safe_rows(
            MenuItem.objects.all(),
            [
                'id', 'category_id', 'name_es', 'name_en',
                'description_es', 'description_en', 'price',
                'image', 'is_active', 'is_available', 'sort_order',
                'created_at', 'updated_at',
            ],
        ),
        'menu_option_groups': _json_safe_rows(
            MenuOptionGroup.objects.all(),
            [
                'id', 'menu_item_id', 'title_es', 'title_en',
                'required', 'min_choices', 'max_choices', 'sort_order',
            ],
        ),
        'menu_options': _json_safe_rows(
            MenuOption.objects.all(),
            [
                'id', 'group_id', 'name_es', 'name_en',
                'extra_price', 'is_active', 'sort_order',
            ],
        ),
        'customers': _json_safe_rows(
            Customer.objects.all(),
            [
                'id', 'name', 'phone', 'email', 'default_address',
                'total_orders', 'last_order_at', 'last_login_at', 'created_at',
            ],
        ),
        'customer_addresses': _json_safe_rows(
            CustomerAddress.objects.all(),
            [
                'id', 'customer_id', 'address_text', 'city', 'postal_code',
                'latitude', 'longitude', 'is_default', 'created_at',
            ],
        ),
        'orders': _json_safe_rows(
            Order.objects.all(),
            [
                'id', 'order_code', 'customer_id', 'customer_name',
                'phone', 'address', 'floor', 'note', 'delivery_type',
                'payment_method', 'status', 'subtotal', 'delivery_fee',
                'discount_amount', 'total', 'assigned_rider_id',
                'delivery_latitude', 'delivery_longitude',
                'route_distance_km', 'route_duration_min',
                'created_at', 'updated_at',
            ],
        ),
        'order_items': _json_safe_rows(
            OrderItem.objects.all(),
            [
                'id', 'order_id', 'menu_item_id', 'name', 'quantity',
                'unit_price', 'options_text', 'total',
            ],
        ),
        'payments': _json_safe_rows(
            Payment.objects.all(),
            [
                'id', 'order_id', 'method', 'status', 'amount',
                'transaction_id', 'created_at', 'updated_at',
            ],
        ),
        'riders': _json_safe_rows(
            Rider.objects.all(),
            [
                'id', 'name', 'phone', 'username', 'is_active',
                'current_latitude', 'current_longitude',
                'last_location_at', 'created_at',
            ],
        ),
        'restaurant_settings': _json_safe_rows(
            RestaurantSettings.objects.all(),
            [
                field.name for field in RestaurantSettings._meta.fields
                if field.name not in {'id'}
            ],
        ),
        'coupons': _json_safe_rows(
            Coupon.objects.all(),
            [field.name for field in Coupon._meta.fields],
        ),
        'expense_categories': _json_safe_rows(
            ExpenseCategory.objects.all(),
            ['id', 'name', 'is_active', 'sort_order', 'created_at'],
        ),
        'accounting_settings': _json_safe_rows(
            AccountingSettings.objects.all(),
            [
                'id', 'saeid_share_percent', 'ahmed_share_percent',
                'bbva_initial_balance', 'updated_at',
            ],
        ),
        'financial_entries': _json_safe_rows(
            RestaurantFinancialEntry.objects.all(),
            [
                'id', 'entry_type', 'title', 'description', 'amount',
                'entry_date', 'category_id', 'paid_by', 'contribution_from',
                'settlement_to', 'payment_method', 'invoice_number',
                'bank_reference', 'receipt', 'status',
                'created_by_username', 'updated_by_username',
                'created_at', 'updated_at',
            ],
        ),
        'reviews': _json_safe_rows(
            OrderReview.objects.all(),
            [field.name for field in OrderReview._meta.fields],
        ),
        'order_chat_messages': _json_safe_rows(
            OrderChatMessage.objects.all(),
            [field.name for field in OrderChatMessage._meta.fields],
        ),
    }


def _configuration_backup_payload():
    from .models import ExpenseCategory, AccountingSettings

    return {
        'metadata': {
            'project': 'Casa de Kebab Turco',
            'created_at': timezone.now(),
            'format_version': 1,
        },
        'restaurant_settings': _json_safe_rows(
            RestaurantSettings.objects.all(),
            [field.name for field in RestaurantSettings._meta.fields],
        ),
        'categories': _json_safe_rows(
            Category.objects.all(),
            ['id', 'name_es', 'name_en', 'slug', 'sort_order', 'is_active'],
        ),
        'coupons': _json_safe_rows(
            Coupon.objects.all(),
            [field.name for field in Coupon._meta.fields],
        ),
        'expense_categories': _json_safe_rows(
            ExpenseCategory.objects.all(),
            ['id', 'name', 'is_active', 'sort_order'],
        ),
        'accounting_settings': _json_safe_rows(
            AccountingSettings.objects.all(),
            [
                'id', 'saeid_share_percent', 'ahmed_share_percent',
                'bbva_initial_balance',
            ],
        ),
    }


def _create_json_backup(backup, payload):
    timestamp = timezone.localtime().strftime('%Y-%m-%d_%H-%M-%S')
    file_name = f'casa_kebab_{backup.backup_type}_{timestamp}.json'
    file_path = os.path.join(_backup_root(), file_name)

    with open(file_path, 'w', encoding='utf-8') as stream:
        json.dump(payload, stream, cls=DjangoJSONEncoder, ensure_ascii=False, indent=2)

    return file_name, file_path


def _create_media_backup(backup):
    timestamp = timezone.localtime().strftime('%Y-%m-%d_%H-%M-%S')
    file_name = f'casa_kebab_media_{timestamp}.zip'
    file_path = os.path.join(_backup_root(), file_name)

    media_root = os.path.abspath(str(settings.MEDIA_ROOT))
    backup_root = os.path.abspath(_backup_root())

    with zipfile.ZipFile(file_path, 'w', zipfile.ZIP_DEFLATED) as archive:
        if os.path.isdir(media_root):
            for root, _, files in os.walk(media_root):
                root_abs = os.path.abspath(root)
                if root_abs.startswith(backup_root):
                    continue
                for file_name_item in files:
                    absolute_path = os.path.join(root, file_name_item)
                    relative_path = os.path.relpath(absolute_path, media_root)
                    archive.write(absolute_path, relative_path)

    return file_name, file_path


@api_view(['GET'])
@admin_token_required
def admin_system_health(request):
    health = {
        'backend': {'status': 'ok', 'label': 'Operativo', 'detail': 'Django API responde correctamente.'},
        'database': {'status': 'unknown', 'label': 'Comprobando', 'detail': ''},
        'media': {'status': 'unknown', 'label': 'Comprobando', 'detail': ''},
        'sms_gateway': {
            'status': 'ok' if getattr(settings, 'SMS_GATEWAY_TOKEN', '') else 'warning',
            'label': 'Configurado' if getattr(settings, 'SMS_GATEWAY_TOKEN', '') else 'Sin configurar',
            'detail': 'Token SMS disponible.' if getattr(settings, 'SMS_GATEWAY_TOKEN', '') else 'No hay token SMS configurado.',
        },
        'telegram': {
            'status': 'ok' if (
                getattr(settings, 'TELEGRAM_ENABLED', False)
                and getattr(settings, 'TELEGRAM_BOT_TOKEN', '')
            ) else 'warning',
            'label': 'Configurado' if (
                getattr(settings, 'TELEGRAM_ENABLED', False)
                and getattr(settings, 'TELEGRAM_BOT_TOKEN', '')
            ) else 'Desactivado',
            'detail': 'Telegram está habilitado.' if getattr(settings, 'TELEGRAM_ENABLED', False) else 'Telegram está desactivado.',
        },
    }

    try:
        with connection.cursor() as cursor:
            cursor.execute('SELECT 1')
            cursor.fetchone()
        health['database'] = {
            'status': 'ok',
            'label': 'Operativo',
            'detail': connection.vendor,
        }
    except Exception as exc:
        health['database'] = {
            'status': 'error',
            'label': 'No disponible',
            'detail': str(exc),
        }

    try:
        root = _backup_root()
        test_path = os.path.join(root, '.write_test')
        with open(test_path, 'w', encoding='utf-8') as stream:
            stream.write('ok')
        os.remove(test_path)
        health['media'] = {
            'status': 'ok',
            'label': 'Escritura disponible',
            'detail': str(settings.MEDIA_ROOT),
        }
    except Exception as exc:
        health['media'] = {
            'status': 'error',
            'label': 'No escribible',
            'detail': str(exc),
        }

    latest = SystemBackup.objects.filter(
        status=SystemBackup.STATUS_COMPLETED
    ).order_by('-completed_at').first()

    return Response({
        'health': health,
        'latest_backup': SystemBackupSerializer(latest).data if latest else None,
        'backup_count': SystemBackup.objects.count(),
        'completed_count': SystemBackup.objects.filter(
            status=SystemBackup.STATUS_COMPLETED
        ).count(),
        'failed_count': SystemBackup.objects.filter(
            status=SystemBackup.STATUS_FAILED
        ).count(),
        'render_warning': (
            'Render local filesystem may be ephemeral. '
            'Download backups immediately or configure Persistent Disk/S3.'
        ),
    })


@api_view(['GET', 'POST'])
@admin_token_required
def admin_system_backups(request):
    if request.method == 'GET':
        qs = SystemBackup.objects.all()[:100]
        return Response(SystemBackupSerializer(qs, many=True).data)

    backup_type = (request.data.get('backup_type') or '').strip()
    allowed_types = {
        SystemBackup.TYPE_DATABASE,
        SystemBackup.TYPE_CONFIGURATION,
        SystemBackup.TYPE_MEDIA,
    }
    if backup_type not in allowed_types:
        return Response(
            {'detail': 'Tipo de copia no válido.'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    backup = SystemBackup.objects.create(
        backup_type=backup_type,
        status=SystemBackup.STATUS_RUNNING,
        created_by_username=request.admin_user.get_username(),
    )

    try:
        if backup_type == SystemBackup.TYPE_DATABASE:
            file_name, file_path = _create_json_backup(
                backup,
                _database_backup_payload(),
            )
        elif backup_type == SystemBackup.TYPE_CONFIGURATION:
            file_name, file_path = _create_json_backup(
                backup,
                _configuration_backup_payload(),
            )
        else:
            file_name, file_path = _create_media_backup(backup)

        backup.file_name = file_name
        backup.file_path = file_path
        backup.file_size = os.path.getsize(file_path)
        backup.checksum_sha256 = _sha256_file(file_path)
        backup.status = SystemBackup.STATUS_COMPLETED
        backup.completed_at = timezone.now()
        backup.save(update_fields=[
            'file_name', 'file_path', 'file_size',
            'checksum_sha256', 'status', 'completed_at',
        ])
    except Exception as exc:
        backup.status = SystemBackup.STATUS_FAILED
        backup.error_message = str(exc)
        backup.completed_at = timezone.now()
        backup.save(update_fields=[
            'status', 'error_message', 'completed_at',
        ])
        return Response(
            {
                'detail': 'No se pudo crear la copia.',
                'backup': SystemBackupSerializer(backup).data,
            },
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

    return Response(
        SystemBackupSerializer(backup).data,
        status=status.HTTP_201_CREATED,
    )


@api_view(['GET', 'PATCH', 'DELETE'])
@admin_token_required
def admin_system_backup_detail(request, backup_id):
    try:
        backup = SystemBackup.objects.get(id=backup_id)
    except SystemBackup.DoesNotExist:
        return Response(
            {'detail': 'Copia no encontrada.'},
            status=status.HTTP_404_NOT_FOUND,
        )

    if request.method == 'GET':
        return Response(SystemBackupSerializer(backup).data)

    if request.method == 'PATCH':
        if 'is_protected' not in request.data:
            return Response(
                {'detail': 'is_protected es obligatorio.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        backup.is_protected = bool(request.data.get('is_protected'))
        backup.save(update_fields=['is_protected'])
        return Response(SystemBackupSerializer(backup).data)

    if backup.is_protected:
        return Response(
            {'detail': 'La copia está protegida. Desprotégela antes de eliminarla.'},
            status=status.HTTP_409_CONFLICT,
        )

    if backup.file_path and os.path.isfile(backup.file_path):
        try:
            os.remove(backup.file_path)
        except OSError:
            pass
    backup.delete()
    return Response(status=status.HTTP_204_NO_CONTENT)


@api_view(['GET'])
@admin_token_required
def admin_system_backup_download(request, backup_id):
    try:
        backup = SystemBackup.objects.get(
            id=backup_id,
            status=SystemBackup.STATUS_COMPLETED,
        )
    except SystemBackup.DoesNotExist:
        return Response(
            {'detail': 'Copia no encontrada o incompleta.'},
            status=status.HTTP_404_NOT_FOUND,
        )

    if not backup.file_path or not os.path.isfile(backup.file_path):
        return Response(
            {'detail': 'El archivo ya no existe en el servidor.'},
            status=status.HTTP_410_GONE,
        )

    return FileResponse(
        open(backup.file_path, 'rb'),
        as_attachment=True,
        filename=backup.file_name,
    )


@api_view(['POST'])
@admin_token_required
def admin_system_backup_verify(request, backup_id):
    try:
        backup = SystemBackup.objects.get(id=backup_id)
    except SystemBackup.DoesNotExist:
        return Response(
            {'detail': 'Copia no encontrada.'},
            status=status.HTTP_404_NOT_FOUND,
        )

    if not backup.file_path or not os.path.isfile(backup.file_path):
        return Response({
            'valid': False,
            'detail': 'El archivo no existe en el servidor.',
        }, status=status.HTTP_410_GONE)

    actual_checksum = _sha256_file(backup.file_path)
    valid = bool(
        backup.checksum_sha256
        and actual_checksum == backup.checksum_sha256
    )
    return Response({
        'valid': valid,
        'expected_checksum': backup.checksum_sha256,
        'actual_checksum': actual_checksum,
        'detail': (
            'La copia es válida.'
            if valid
            else 'La suma de verificación no coincide.'
        ),
    })

# --- Rider FCM endpoints restored by targeted repair ---
@api_view(["POST"])
@rider_token_required
def secure_rider_push_register(request):
    from django.utils import timezone
    from .models import RiderPushDevice
    from .serializers import RiderPushDeviceSerializer

    rider = request.rider
    token = str(
        request.data.get("device_token")
        or request.data.get("token")
        or ""
    ).strip()

    if not token:
        return Response(
            {"detail": "device_token is required"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    field_names = {
        field.name for field in RiderPushDevice._meta.get_fields()
        if getattr(field, "concrete", False)
    }

    lookup_name = (
        "device_token" if "device_token" in field_names
        else "token" if "token" in field_names
        else None
    )
    if not lookup_name:
        return Response(
            {"detail": "RiderPushDevice has no token field"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

    defaults = {}
    if "rider" in field_names:
        defaults["rider"] = rider
    if "platform" in field_names:
        defaults["platform"] = str(
            request.data.get("platform") or "android"
        ).strip().lower()
    if "app_version" in field_names:
        defaults["app_version"] = str(
            request.data.get("app_version") or ""
        ).strip()[:80]
    if "is_active" in field_names:
        defaults["is_active"] = True
    if "last_seen_at" in field_names:
        defaults["last_seen_at"] = timezone.now()
    if "last_error" in field_names:
        defaults["last_error"] = ""

    device, created = RiderPushDevice.objects.update_or_create(
        **{lookup_name: token},
        defaults=defaults,
    )

    return Response(
        {
            "success": True,
            "created": created,
            "device": RiderPushDeviceSerializer(device).data,
        },
        status=(
            status.HTTP_201_CREATED
            if created
            else status.HTTP_200_OK
        ),
    )


@api_view(["POST"])
@rider_token_required
def secure_rider_push_unregister(request):
    from django.utils import timezone
    from .models import RiderPushDevice

    rider = request.rider
    token = str(
        request.data.get("device_token")
        or request.data.get("token")
        or ""
    ).strip()

    if not token:
        return Response(
            {"detail": "device_token is required"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    field_names = {
        field.name for field in RiderPushDevice._meta.get_fields()
        if getattr(field, "concrete", False)
    }
    lookup_name = (
        "device_token" if "device_token" in field_names
        else "token" if "token" in field_names
        else None
    )
    if not lookup_name:
        return Response(
            {"detail": "RiderPushDevice has no token field"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

    filters = {lookup_name: token}
    if "rider" in field_names:
        filters["rider"] = rider

    updates = {}
    if "is_active" in field_names:
        updates["is_active"] = False
    if "last_seen_at" in field_names:
        updates["last_seen_at"] = timezone.now()

    updated = RiderPushDevice.objects.filter(**filters).update(**updates)
    return Response({"success": True, "updated": updated})


@api_view(["POST"])
@rider_token_required
def secure_rider_push_test(request):
    try:
        from .rider_push_notifications import send_new_order_to_rider
    except Exception as exc:
        return Response(
            {"detail": f"Push helper unavailable: {exc}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

    rider = request.rider
    order = None

    try:
        order = rider.orders.exclude(
            status__in=["delivered", "cancelled"]
        ).order_by("-created_at").first()
    except Exception:
        pass

    if not order:
        return Response(
            {"detail": "No active rider order is available for a test."},
            status=status.HTTP_409_CONFLICT,
        )

    try:
        result = send_new_order_to_rider(order)
    except Exception as exc:
        return Response(
            {"detail": f"Push test failed: {exc}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )

    return Response({"success": True, "result": result})
# --- End Rider FCM endpoint repair ---

@api_view(['GET'])
def admin_dynamic_reports(request):
    admin_user = get_admin_user_from_request(request)
    if not admin_user:
        return Response({'detail': 'Admin authentication is required.'}, status=status.HTTP_401_UNAUTHORIZED)

    from datetime import timedelta
    from decimal import Decimal
    from django.db.models import Count, Sum, F, DecimalField, ExpressionWrapper, Max, Q
    from django.db.models.functions import TruncDate, ExtractHour
    from django.utils.dateparse import parse_date
    from .models import Order, OrderItem

    today = timezone.localdate()
    date_from = parse_date(str(request.query_params.get('date_from') or '')) or (today - timedelta(days=29))
    date_to = parse_date(str(request.query_params.get('date_to') or '')) or today
    if date_from > date_to:
        return Response({'detail': 'La fecha inicial no puede ser posterior a la fecha final.'}, status=status.HTTP_400_BAD_REQUEST)

    orders = Order.objects.filter(created_at__date__gte=date_from, created_at__date__lte=date_to)
    delivery_type = str(request.query_params.get('delivery_type') or '').strip()
    payment_method = str(request.query_params.get('payment_method') or '').strip()
    status_filter = str(request.query_params.get('status') or '').strip()
    rider_id = str(request.query_params.get('rider_id') or '').strip()
    if delivery_type: orders = orders.filter(delivery_type=delivery_type)
    if payment_method: orders = orders.filter(payment_method=payment_method)
    if status_filter: orders = orders.filter(status=status_filter)
    if rider_id.isdigit(): orders = orders.filter(assigned_rider_id=int(rider_id))

    cancelled_qs = orders.filter(status=Order.STATUS_CANCELLED)
    valid_orders = orders.exclude(status=Order.STATUS_CANCELLED)
    revenue = valid_orders.aggregate(value=Sum('total'))['value'] or Decimal('0.00')
    discounts = valid_orders.aggregate(value=Sum('discount'))['value'] or Decimal('0.00')
    orders_count = valid_orders.count()
    cancelled_count = cancelled_qs.count()
    average_order = (revenue / orders_count) if orders_count else Decimal('0.00')
    item_revenue = ExpressionWrapper(F('price_snapshot') * F('quantity'), output_field=DecimalField(max_digits=12, decimal_places=2))

    daily_sales = [{'day': row['day'].isoformat() if row['day'] else '', 'orders': row['orders'] or 0, 'revenue': float(row['revenue'] or 0)} for row in valid_orders.annotate(day=TruncDate('created_at')).values('day').annotate(orders=Count('id'), revenue=Sum('total')).order_by('day')]
    top_items = [{'name': row['name_snapshot'] or 'Producto', 'quantity': int(row['quantity'] or 0), 'revenue': float(row['revenue'] or 0)} for row in OrderItem.objects.filter(order__in=valid_orders).values('name_snapshot').annotate(quantity=Sum('quantity'), revenue=Sum(item_revenue)).order_by('-quantity', '-revenue')[:10]]

    status_labels = {'pending':'Pendiente','accepted':'Aceptado','preparing':'Preparando','out_for_delivery':'En reparto','delivered':'Entregado','cancelled':'Cancelado'}
    status_breakdown = [{'label': status_labels.get(row['status'], row['status']), 'count': int(row['count'] or 0), 'revenue': float(row['revenue'] or 0)} for row in orders.values('status').annotate(count=Count('id'), revenue=Sum('total')).order_by('-count')]
    payment_labels = {'cash':'Efectivo','card_delivery':'Tarjeta','store':'En tienda','online':'Online'}
    payment_breakdown = [{'label': payment_labels.get(row['payment_method'], row['payment_method']), 'count': int(row['count'] or 0), 'revenue': float(row['revenue'] or 0)} for row in valid_orders.values('payment_method').annotate(count=Count('id'), revenue=Sum('total')).order_by('-revenue')]
    hourly_sales = [{'hour': f"{int(row['hour'] or 0):02d}:00", 'orders': int(row['orders'] or 0), 'revenue': float(row['revenue'] or 0)} for row in valid_orders.annotate(hour=ExtractHour('created_at')).values('hour').annotate(orders=Count('id'), revenue=Sum('total')).order_by('hour')]
    rider_performance = [{'id':row['assigned_rider_id'], 'name':row['assigned_rider__name'] or 'Repartidor', 'orders':int(row['orders'] or 0), 'delivered':int(row['delivered'] or 0), 'revenue':float(row['revenue'] or 0)} for row in valid_orders.filter(assigned_rider__isnull=False).values('assigned_rider_id','assigned_rider__name').annotate(orders=Count('id'), delivered=Count('id', filter=Q(status=Order.STATUS_DELIVERED)), revenue=Sum('total')).order_by('-delivered','-orders')]
    top_customers = [{'name':row['customer_name'] or 'Sin nombre','phone':row['customer_phone'] or '','orders':int(row['orders'] or 0),'revenue':float(row['revenue'] or 0),'last_order':row['last_order'].isoformat() if row['last_order'] else None} for row in valid_orders.exclude(customer_phone='').values('customer_name','customer_phone').annotate(orders=Count('id'), revenue=Sum('total'), last_order=Max('created_at')).order_by('-revenue','-orders')[:20]]

    return Response({
        'filters': {'date_from':date_from.isoformat(),'date_to':date_to.isoformat(),'delivery_type':delivery_type,'payment_method':payment_method,'status':status_filter,'rider_id':rider_id},
        'metrics': {'orders_count':orders_count,'revenue':float(revenue),'average_order':float(average_order),'delivery_orders':valid_orders.filter(delivery_type=Order.DELIVERY_DELIVERY).count(),'collection_orders':valid_orders.filter(delivery_type=Order.DELIVERY_COLLECTION).count(),'cancelled_orders':cancelled_count,'cancel_rate':round((cancelled_count / orders.count() * 100), 1) if orders.exists() else 0,'unique_customers':valid_orders.exclude(customer_phone='').values('customer_phone').distinct().count(),'discount_total':float(discounts)},
        'daily_sales':daily_sales,'top_items':top_items,'status_breakdown':status_breakdown,'payment_breakdown':payment_breakdown,'hourly_sales':hourly_sales,'rider_performance':rider_performance,'top_customers':top_customers,
    })

# v18 profitability endpoints — restricted to Admin
def _profitability_ingredient_payload(ingredient):
    return {
        'id': ingredient.id,
        'name': ingredient.name,
        'unit': ingredient.unit,
        'unit_cost': float(ingredient.unit_cost or 0),
        'stock_quantity': float(ingredient.stock_quantity or 0),
        'reorder_level': float(ingredient.reorder_level or 0),
        'supplier_name': ingredient.supplier_name or '',
        'is_active': bool(ingredient.is_active),
        'updated_at': ingredient.updated_at.isoformat() if ingredient.updated_at else None,
    }


def _profitability_profile_payload(profile):
    components = []
    for component in profile.components.select_related('ingredient').all():
        components.append({
            'id': component.id,
            'ingredient_id': component.ingredient_id,
            'ingredient_name': component.ingredient.name,
            'unit': component.ingredient.unit,
            'unit_cost': float(component.ingredient.unit_cost or 0),
            'quantity': float(component.quantity or 0),
            'line_cost': float(component.line_cost or 0),
        })

    ingredient_cost = sum(Decimal(str(row['line_cost'])) for row in components)
    total_cost = ingredient_cost + (profile.packaging_cost or Decimal('0.00')) + (profile.fixed_cost or Decimal('0.00'))
    price = profile.menu_item.price or Decimal('0.00')
    gross_profit = price - total_cost
    margin = (gross_profit / price * Decimal('100.00')) if price > 0 else Decimal('0.00')

    return {
        'menu_item_id': profile.menu_item_id,
        'menu_item_name': profile.menu_item.name_es,
        'selling_price': float(price),
        'packaging_cost': float(profile.packaging_cost or 0),
        'fixed_cost': float(profile.fixed_cost or 0),
        'target_margin_percent': float(profile.target_margin_percent or 0),
        'notes': profile.notes or '',
        'ingredient_cost': float(ingredient_cost),
        'total_unit_cost': float(total_cost),
        'gross_profit_per_unit': float(gross_profit),
        'margin_percent': float(margin),
        'components': components,
    }


@api_view(['GET', 'POST'])
@admin_token_required
def admin_profitability_ingredients(request):
    

    if request.method == 'GET':
        qs = Ingredient.objects.all().order_by('name')
        return Response([_profitability_ingredient_payload(row) for row in qs])

    data = request.data or {}
    name = str(data.get('name') or '').strip()
    if not name:
        return Response({'detail': 'El nombre del ingrediente es obligatorio.'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        ingredient = Ingredient.objects.create(
            name=name,
            unit=str(data.get('unit') or 'g'),
            unit_cost=Decimal(str(data.get('unit_cost') or '0')),
            stock_quantity=Decimal(str(data.get('stock_quantity') or '0')),
            reorder_level=Decimal(str(data.get('reorder_level') or '0')),
            supplier_name=str(data.get('supplier_name') or '').strip(),
            is_active=bool(data.get('is_active', True)),
        )
    except Exception as exc:
        return Response({'detail': f'No se pudo crear el ingrediente: {exc}'}, status=status.HTTP_400_BAD_REQUEST)
    return Response(_profitability_ingredient_payload(ingredient), status=status.HTTP_201_CREATED)


@api_view(['PATCH', 'DELETE'])
@admin_token_required
def admin_profitability_ingredient_detail(request, ingredient_id):
    

    try:
        ingredient = Ingredient.objects.get(id=ingredient_id)
    except Ingredient.DoesNotExist:
        return Response({'detail': 'Ingrediente no encontrado.'}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'DELETE':
        if ingredient.recipe_components.exists():
            return Response({'detail': 'No se puede eliminar porque este ingrediente ya está usado en recetas. Desactívalo o elimina las recetas relacionadas.'}, status=status.HTTP_400_BAD_REQUEST)
        ingredient.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    data = request.data or {}
    allowed = ['name', 'unit', 'supplier_name', 'is_active']
    for field in allowed:
        if field in data:
            setattr(ingredient, field, str(data[field]).strip() if field != 'is_active' else bool(data[field]))
    for field in ['unit_cost', 'stock_quantity', 'reorder_level']:
        if field in data:
            setattr(ingredient, field, Decimal(str(data[field] or '0')))
    try:
        ingredient.full_clean()
        ingredient.save()
    except Exception as exc:
        return Response({'detail': f'No se pudo actualizar el ingrediente: {exc}'}, status=status.HTTP_400_BAD_REQUEST)
    return Response(_profitability_ingredient_payload(ingredient))


@api_view(['GET', 'PUT'])
@admin_token_required
def admin_profitability_recipe(request, menu_item_id):
    from .models import ProductCostProfile

    try:
        item = MenuItem.objects.get(id=menu_item_id)
    except MenuItem.DoesNotExist:
        return Response({'detail': 'Producto de menú no encontrado.'}, status=status.HTTP_404_NOT_FOUND)

    profile, _ = ProductCostProfile.objects.get_or_create(menu_item=item)

    if request.method == 'GET':
        return Response(_profitability_profile_payload(profile))

    data = request.data or {}
    raw_components = data.get('components', [])
    if not isinstance(raw_components, list):
        return Response({'detail': 'components debe ser una lista.'}, status=status.HTTP_400_BAD_REQUEST)

    normalized = []
    used_ids = set()
    try:
        for row in raw_components:
            ingredient_id = int(row.get('ingredient_id'))
            quantity = Decimal(str(row.get('quantity') or '0'))
            if quantity <= 0:
                raise ValueError('La cantidad de cada ingrediente debe ser mayor que cero.')
            if ingredient_id in used_ids:
                raise ValueError('Un ingrediente no puede repetirse en la misma receta.')
            ingredient = Ingredient.objects.get(id=ingredient_id, is_active=True)
            used_ids.add(ingredient_id)
            normalized.append((ingredient, quantity))
    except (ValueError, TypeError, Ingredient.DoesNotExist) as exc:
        return Response({'detail': f'Receta inválida: {exc}'}, status=status.HTTP_400_BAD_REQUEST)

    try:
        with transaction.atomic():
            profile.packaging_cost = Decimal(str(data.get('packaging_cost') or '0'))
            profile.fixed_cost = Decimal(str(data.get('fixed_cost') or '0'))
            profile.target_margin_percent = Decimal(str(data.get('target_margin_percent') or '55'))
            profile.notes = str(data.get('notes') or '').strip()
            profile.full_clean()
            profile.save()
            profile.components.all().delete()
            RecipeIngredient.objects.bulk_create([
                RecipeIngredient(profile=profile, ingredient=ingredient, quantity=quantity)
                for ingredient, quantity in normalized
            ])
    except Exception as exc:
        return Response({'detail': f'No se pudo guardar la receta: {exc}'}, status=status.HTTP_400_BAD_REQUEST)

    profile.refresh_from_db()
    return Response(_profitability_profile_payload(profile))


@api_view(['GET'])
@admin_token_required
def admin_profitability_report(request):
    from datetime import timedelta
    from django.utils.dateparse import parse_date
    from django.db.models import Sum
    from .models import ProductCostProfile, OrderItem

    today = timezone.localdate()
    date_from = parse_date(str(request.query_params.get('date_from') or '')) or (today - timedelta(days=29))
    date_to = parse_date(str(request.query_params.get('date_to') or '')) or today
    if date_from > date_to:
        return Response({'detail': 'La fecha inicial no puede ser posterior a la final.'}, status=status.HTTP_400_BAD_REQUEST)

    sales_rows = (
        OrderItem.objects.filter(
            order__created_at__date__gte=date_from,
            order__created_at__date__lte=date_to,
        )
        .exclude(order__status=Order.STATUS_CANCELLED)
        .values('menu_item_id')
        .annotate(units_sold=Sum('quantity'), sales_revenue=Sum('total'))
    )
    sales_by_item = {
        row['menu_item_id']: {
            'units_sold': int(row['units_sold'] or 0),
            'sales_revenue': Decimal(row['sales_revenue'] or '0.00'),
        }
        for row in sales_rows if row['menu_item_id']
    }

    result = []
    for profile in ProductCostProfile.objects.select_related('menu_item').prefetch_related('components__ingredient').all():
        payload = _profitability_profile_payload(profile)
        sales = sales_by_item.get(profile.menu_item_id, {'units_sold': 0, 'sales_revenue': Decimal('0.00')})
        total_cost = Decimal(str(payload['total_unit_cost']))
        units = Decimal(str(sales['units_sold']))
        payload.update({
            'units_sold': int(sales['units_sold']),
            'sales_revenue': float(sales['sales_revenue']),
            'estimated_cost_of_sales': float(total_cost * units),
            'estimated_gross_profit': float((Decimal(str(payload['gross_profit_per_unit'])) * units)),
            'has_recipe': bool(payload['components']),
        })
        result.append(payload)

    result.sort(key=lambda row: (-row['estimated_gross_profit'], -row['units_sold'], row['menu_item_name']))
    warning_items = [
        row for row in result
        if row['has_recipe'] and row['margin_percent'] < row['target_margin_percent']
    ]

    return Response({
        'date_from': date_from.isoformat(),
        'date_to': date_to.isoformat(),
        'items': result,
        'warnings': [{
            'menu_item_id': row['menu_item_id'],
            'menu_item_name': row['menu_item_name'],
            'margin_percent': row['margin_percent'],
            'target_margin_percent': row['target_margin_percent'],
        } for row in warning_items],
        'summary': {
            'configured_products': len(result),
            'products_below_target': len(warning_items),
            'sales_revenue': float(sum(Decimal(str(row['sales_revenue'])) for row in result)),
            'estimated_cost_of_sales': float(sum(Decimal(str(row['estimated_cost_of_sales'])) for row in result)),
            'estimated_gross_profit': float(sum(Decimal(str(row['estimated_gross_profit'])) for row in result)),
        },
    })

@api_view(['GET'])
def public_customer_menu_highlights(request):
    """
    Customer-facing menu highlights based on valid recent orders.
    It only returns statistics and does not edit categories, prices, or menu data.
    """
    from datetime import timedelta
    from django.db.models import Sum

    try:
        days = int(request.query_params.get('days', 30))
    except (TypeError, ValueError):
        days = 30
    days = max(7, min(days, 180))
    start_date = timezone.localdate() - timedelta(days=days - 1)

    rows = (
        OrderItem.objects.filter(order__created_at__date__gte=start_date)
        .exclude(order__status=Order.STATUS_CANCELLED)
        .values('menu_item_id')
        .annotate(units_sold=Sum('quantity'))
    )
    sales = {
        str(row['menu_item_id']): int(row['units_sold'] or 0)
        for row in rows
        if row.get('menu_item_id')
    }

    available_items = MenuItem.objects.filter(is_active=True, is_available=True)
    cheapest = available_items.order_by('price', 'sort_order', 'id').first()

    top_seller_id = None
    top_units = 0
    for item_id, units in sales.items():
        if units > top_units:
            top_seller_id = item_id
            top_units = units

    # If no completed order exists yet, use the first active product as a safe fallback.
    if not top_seller_id:
        first_item = available_items.order_by('sort_order', 'id').first()
        top_seller_id = str(first_item.id) if first_item else None

    return Response({
        'days': days,
        'sales_by_item': sales,
        'top_seller_id': top_seller_id,
        'top_seller_units': top_units,
        'lowest_price_item_id': str(cheapest.id) if cheapest else None,
    })

# ============================================================
# Finance Reports v2 - Casa de Kebab Turco
# Uses existing models only. No new database tables.
# ============================================================

from decimal import Decimal
from django.db.models import Sum, Count
from django.utils import timezone
from django.utils.dateparse import parse_date
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAdminUser
from rest_framework.response import Response

from .models import (
    Order,
    OrderItem,
    RestaurantFinancialEntry,
    AccountingSettings,
)


def _finance_v2_money(value):
    value = Decimal(str(value or "0.00"))
    return str(value.quantize(Decimal("0.01")))


def _finance_v2_date_range(request):
    start = parse_date(request.GET.get("start", ""))
    end = parse_date(request.GET.get("end", ""))
    if not start or not end:
        today = timezone.localdate()
        start = today.replace(day=1)
        end = today
    return start, end


def _finance_v2_orders_between(start, end):
    return Order.objects.filter(
        created_at__date__gte=start,
        created_at__date__lte=end,
    ).exclude(status=Order.STATUS_CANCELLED)


def _finance_v2_entries_between(start, end):
    return RestaurantFinancialEntry.objects.filter(
        entry_date__gte=start,
        entry_date__lte=end,
        status=RestaurantFinancialEntry.STATUS_APPROVED,
    )


@api_view(["GET"])
@permission_classes([IsAdminUser])
def finance_profit_loss_v2(request):
    start, end = _finance_v2_date_range(request)

    orders = _finance_v2_orders_between(start, end)
    entries = _finance_v2_entries_between(start, end)

    revenue = orders.aggregate(s=Sum("total"))["s"] or Decimal("0.00")
    subtotal = orders.aggregate(s=Sum("subtotal"))["s"] or Decimal("0.00")
    delivery_fees = orders.aggregate(s=Sum("delivery_fee"))["s"] or Decimal("0.00")
    discounts = orders.aggregate(s=Sum("discount"))["s"] or Decimal("0.00")

    expenses = entries.filter(
        entry_type=RestaurantFinancialEntry.TYPE_EXPENSE
    ).aggregate(s=Sum("amount"))["s"] or Decimal("0.00")

    contributions = entries.filter(
        entry_type=RestaurantFinancialEntry.TYPE_CONTRIBUTION
    ).aggregate(s=Sum("amount"))["s"] or Decimal("0.00")

    settlements = entries.filter(
        entry_type=RestaurantFinancialEntry.TYPE_SETTLEMENT
    ).aggregate(s=Sum("amount"))["s"] or Decimal("0.00")

    net_profit = revenue - expenses
    orders_count = orders.count()
    avg_order_value = revenue / orders_count if orders_count else Decimal("0.00")

    by_payment_method = list(
        orders.values("payment_method")
        .annotate(count=Count("id"), total=Sum("total"))
        .order_by("-total")
    )

    by_delivery_type = list(
        orders.values("delivery_type")
        .annotate(count=Count("id"), total=Sum("total"))
        .order_by("-total")
    )

    expense_by_category = list(
        entries.filter(entry_type=RestaurantFinancialEntry.TYPE_EXPENSE)
        .values("category__name")
        .annotate(count=Count("id"), total=Sum("amount"))
        .order_by("-total")
    )

    expense_by_paid_by = list(
        entries.filter(entry_type=RestaurantFinancialEntry.TYPE_EXPENSE)
        .values("paid_by")
        .annotate(count=Count("id"), total=Sum("amount"))
        .order_by("-total")
    )

    return Response({
        "period": {"start": start, "end": end},
        "summary": {
            "revenue": _finance_v2_money(revenue),
            "subtotal": _finance_v2_money(subtotal),
            "delivery_fees": _finance_v2_money(delivery_fees),
            "discounts": _finance_v2_money(discounts),
            "expenses": _finance_v2_money(expenses),
            "contributions_to_bbva": _finance_v2_money(contributions),
            "settlements": _finance_v2_money(settlements),
            "net_profit": _finance_v2_money(net_profit),
            "orders_count": orders_count,
            "average_order_value": _finance_v2_money(avg_order_value),
        },
        "orders_by_payment_method": [
            {"payment_method": x["payment_method"], "count": x["count"], "total": _finance_v2_money(x["total"])}
            for x in by_payment_method
        ],
        "orders_by_delivery_type": [
            {"delivery_type": x["delivery_type"], "count": x["count"], "total": _finance_v2_money(x["total"])}
            for x in by_delivery_type
        ],
        "expenses_by_category": [
            {"category": x["category__name"] or "Sin categoría", "count": x["count"], "total": _finance_v2_money(x["total"])}
            for x in expense_by_category
        ],
        "expenses_by_paid_by": [
            {"paid_by": x["paid_by"], "count": x["count"], "total": _finance_v2_money(x["total"])}
            for x in expense_by_paid_by
        ],
    })


@api_view(["GET"])
@permission_classes([IsAdminUser])
def finance_product_sales_v2(request):
    start, end = _finance_v2_date_range(request)
    orders = _finance_v2_orders_between(start, end)

    rows = (
        OrderItem.objects
        .filter(order__in=orders)
        .values("menu_item_id", "name_snapshot")
        .annotate(quantity_sold=Sum("quantity"), revenue=Sum("total"))
        .order_by("-revenue")
    )

    data = []
    for row in rows:
        revenue = row["revenue"] or Decimal("0.00")
        qty = row["quantity_sold"] or 0
        avg_price = revenue / Decimal(qty) if qty else Decimal("0.00")
        data.append({
            "menu_item_id": row["menu_item_id"],
            "product": row["name_snapshot"],
            "quantity_sold": qty,
            "revenue": _finance_v2_money(revenue),
            "average_price": _finance_v2_money(avg_price),
        })

    return Response({"period": {"start": start, "end": end}, "products": data})


@api_view(["GET"])
@permission_classes([IsAdminUser])
def finance_daily_report_v2(request):
    start, end = _finance_v2_date_range(request)

    orders_by_day = (
        _finance_v2_orders_between(start, end)
        .values("created_at__date")
        .annotate(revenue=Sum("total"), orders_count=Count("id"))
        .order_by("created_at__date")
    )

    expenses_by_day = (
        _finance_v2_entries_between(start, end)
        .filter(entry_type=RestaurantFinancialEntry.TYPE_EXPENSE)
        .values("entry_date")
        .annotate(expenses=Sum("amount"))
        .order_by("entry_date")
    )

    result = {}

    for x in orders_by_day:
        d = str(x["created_at__date"])
        result.setdefault(d, {"date": d, "revenue": Decimal("0.00"), "expenses": Decimal("0.00"), "orders_count": 0})
        result[d]["revenue"] = x["revenue"] or Decimal("0.00")
        result[d]["orders_count"] = x["orders_count"]

    for x in expenses_by_day:
        d = str(x["entry_date"])
        result.setdefault(d, {"date": d, "revenue": Decimal("0.00"), "expenses": Decimal("0.00"), "orders_count": 0})
        result[d]["expenses"] = x["expenses"] or Decimal("0.00")

    rows = []
    for d in sorted(result.keys()):
        row = result[d]
        net = row["revenue"] - row["expenses"]
        rows.append({
            "date": row["date"],
            "revenue": _finance_v2_money(row["revenue"]),
            "expenses": _finance_v2_money(row["expenses"]),
            "net_profit": _finance_v2_money(net),
            "orders_count": row["orders_count"],
        })

    return Response({"period": {"start": start, "end": end}, "days": rows})


@api_view(["GET"])
@permission_classes([IsAdminUser])
def finance_partner_summary_v2(request):
    start, end = _finance_v2_date_range(request)
    settings = AccountingSettings.current()
    entries = _finance_v2_entries_between(start, end)

    expenses_by_partner = {"saeid": Decimal("0.00"), "ahmed": Decimal("0.00"), "bbva": Decimal("0.00")}
    for row in entries.filter(entry_type=RestaurantFinancialEntry.TYPE_EXPENSE).values("paid_by").annotate(total=Sum("amount")):
        expenses_by_partner[row["paid_by"]] = row["total"] or Decimal("0.00")

    contributions_by_partner = {"saeid": Decimal("0.00"), "ahmed": Decimal("0.00"), "bbva": Decimal("0.00")}
    for row in entries.filter(entry_type=RestaurantFinancialEntry.TYPE_CONTRIBUTION).values("contribution_from").annotate(total=Sum("amount")):
        if row["contribution_from"]:
            contributions_by_partner[row["contribution_from"]] = row["total"] or Decimal("0.00")

    total_expenses = sum(expenses_by_partner.values(), Decimal("0.00"))
    saeid_expected = total_expenses * (settings.saeid_share_percent / Decimal("100"))
    ahmed_expected = total_expenses * (settings.ahmed_share_percent / Decimal("100"))

    return Response({
        "period": {"start": start, "end": end},
        "shares": {
            "saeid_percent": _finance_v2_money(settings.saeid_share_percent),
            "ahmed_percent": _finance_v2_money(settings.ahmed_share_percent),
        },
        "expenses_paid": {k: _finance_v2_money(v) for k, v in expenses_by_partner.items()},
        "contributions_to_bbva": {k: _finance_v2_money(v) for k, v in contributions_by_partner.items()},
        "expected_expense_share": {
            "saeid": _finance_v2_money(saeid_expected),
            "ahmed": _finance_v2_money(ahmed_expected),
        },
        "balance_hint": {
            "saeid": _finance_v2_money(expenses_by_partner["saeid"] - saeid_expected),
            "ahmed": _finance_v2_money(expenses_by_partner["ahmed"] - ahmed_expected),
        }
    })


@api_view(["GET"])
@permission_classes([IsAdminUser])
def finance_dashboard_v2(request):
    today = timezone.localdate()
    start = today.replace(day=1)
    end = today

    orders = _finance_v2_orders_between(start, end)
    entries = _finance_v2_entries_between(start, end)

    revenue = orders.aggregate(s=Sum("total"))["s"] or Decimal("0.00")
    expenses = entries.filter(entry_type=RestaurantFinancialEntry.TYPE_EXPENSE).aggregate(s=Sum("amount"))["s"] or Decimal("0.00")
    net_profit = revenue - expenses

    top_products = (
        OrderItem.objects.filter(order__in=orders)
        .values("name_snapshot")
        .annotate(quantity_sold=Sum("quantity"), revenue=Sum("total"))
        .order_by("-revenue")[:5]
    )

    latest_expenses = list(
        entries.filter(entry_type=RestaurantFinancialEntry.TYPE_EXPENSE)
        .order_by("-entry_date", "-created_at")
        .values("entry_date", "title", "amount", "paid_by", "category__name")[:10]
    )

    return Response({
        "period": {"start": start, "end": end},
        "cards": {
            "revenue": _finance_v2_money(revenue),
            "expenses": _finance_v2_money(expenses),
            "net_profit": _finance_v2_money(net_profit),
            "orders_count": orders.count(),
        },
        "top_products": [
            {"product": x["name_snapshot"], "quantity_sold": x["quantity_sold"], "revenue": _finance_v2_money(x["revenue"])}
            for x in top_products
        ],
        "latest_expenses": [
            {
                "date": x["entry_date"],
                "title": x["title"],
                "amount": _finance_v2_money(x["amount"]),
                "paid_by": x["paid_by"],
                "category": x["category__name"] or "Sin categoría",
            }
            for x in latest_expenses
        ],
    })

# ============================================================
# Smart Finance & Inventory - Phase 1
# Uses existing orders, accounting, recipes and ingredient stock.
# ============================================================

def _smart_finance_decimal(value):
    return Decimal(str(value or '0.00'))


def _smart_finance_date_range(request):
    from datetime import timedelta
    from django.utils.dateparse import parse_date

    today = timezone.localdate()
    date_from = parse_date(str(request.query_params.get('date_from') or '')) or today
    date_to = parse_date(str(request.query_params.get('date_to') or '')) or today
    if date_from > date_to:
        raise ValueError('La fecha inicial no puede ser posterior a la fecha final.')
    if (date_to - date_from).days > 366:
        raise ValueError('El periodo máximo permitido es de 366 días.')
    return date_from, date_to


def _smart_finance_recurring_amount(rule, date_from, date_to):
    import calendar
    from datetime import timedelta

    start = max(date_from, rule.start_date)
    end = min(date_to, rule.end_date) if rule.end_date else date_to
    if start > end:
        return Decimal('0.00')

    days = (end - start).days + 1
    amount = _smart_finance_decimal(rule.amount)

    if rule.frequency == rule.FREQUENCY_DAILY:
        return amount * Decimal(days)
    if rule.frequency == rule.FREQUENCY_WEEKLY:
        return (amount / Decimal('7')) * Decimal(days)

    total = Decimal('0.00')
    cursor = start
    while cursor <= end:
        days_in_month = calendar.monthrange(cursor.year, cursor.month)[1]
        month_last_day = cursor.replace(day=days_in_month)
        segment_end = min(month_last_day, end)
        segment_days = (segment_end - cursor).days + 1
        total += (amount / Decimal(days_in_month)) * Decimal(segment_days)
        cursor = segment_end + timedelta(days=1)
    return total


def _smart_finance_rule_payload(rule):
    return {
        'id': rule.id,
        'title': rule.title,
        'amount': float(rule.amount or 0),
        'frequency': rule.frequency,
        'category_id': rule.category_id,
        'category_name': rule.category.name if rule.category_id else '',
        'paid_by': rule.paid_by,
        'start_date': rule.start_date.isoformat() if rule.start_date else '',
        'end_date': rule.end_date.isoformat() if rule.end_date else '',
        'is_active': bool(rule.is_active),
        'notes': rule.notes or '',
    }


def _smart_finance_recipe_costs(date_from, date_to):
    from .models import ProductCostProfile, OrderItem

    sales = (
        OrderItem.objects.filter(
            order__created_at__date__gte=date_from,
            order__created_at__date__lte=date_to,
        )
        .exclude(order__status=Order.STATUS_CANCELLED)
        .values('menu_item_id')
        .annotate(units=Sum('quantity'), revenue=Sum('total'))
    )
    sales_by_item = {
        row['menu_item_id']: {
            'units': _smart_finance_decimal(row['units']),
            'revenue': _smart_finance_decimal(row['revenue']),
        }
        for row in sales if row.get('menu_item_id')
    }

    estimated_cost = Decimal('0.00')
    configured_revenue = Decimal('0.00')
    configured_units = Decimal('0.00')
    profiles = ProductCostProfile.objects.select_related('menu_item').prefetch_related('components__ingredient')
    for profile in profiles:
        sold = sales_by_item.pop(profile.menu_item_id, None)
        if not sold or not profile.components.exists():
            continue
        ingredient_cost = sum(
            (_smart_finance_decimal(component.quantity) * _smart_finance_decimal(component.ingredient.unit_cost))
            for component in profile.components.all()
        )
        unit_cost = ingredient_cost + _smart_finance_decimal(profile.packaging_cost) + _smart_finance_decimal(profile.fixed_cost)
        estimated_cost += unit_cost * sold['units']
        configured_revenue += sold['revenue']
        configured_units += sold['units']

    unconfigured_revenue = sum((row['revenue'] for row in sales_by_item.values()), Decimal('0.00'))
    return {
        'estimated_cost': estimated_cost,
        'configured_revenue': configured_revenue,
        'unconfigured_revenue': unconfigured_revenue,
        'configured_units': configured_units,
    }


def _smart_finance_inventory_payload(today):
    from datetime import timedelta
    from .models import Ingredient, RecipeIngredient, OrderItem

    history_start = today - timedelta(days=13)
    sales = (
        OrderItem.objects.filter(
            order__created_at__date__gte=history_start,
            order__created_at__date__lte=today,
        )
        .exclude(order__status=Order.STATUS_CANCELLED)
        .values('menu_item_id')
        .annotate(units=Sum('quantity'))
    )
    units_by_item = {row['menu_item_id']: _smart_finance_decimal(row['units']) for row in sales if row.get('menu_item_id')}

    daily_usage = {}
    for row in RecipeIngredient.objects.select_related('profile__menu_item', 'ingredient').all():
        sold = units_by_item.get(row.profile.menu_item_id, Decimal('0.00'))
        if sold:
            daily_usage[row.ingredient_id] = daily_usage.get(row.ingredient_id, Decimal('0.00')) + (
                _smart_finance_decimal(row.quantity) * sold / Decimal('14')
            )

    result = []
    for ingredient in Ingredient.objects.filter(is_active=True).order_by('name'):
        stock = _smart_finance_decimal(ingredient.stock_quantity)
        reorder = _smart_finance_decimal(ingredient.reorder_level)
        average = daily_usage.get(ingredient.id, Decimal('0.00'))
        days_left = (stock / average) if average > 0 else None
        low = stock <= reorder if reorder > 0 else False
        urgent = days_left is not None and days_left <= Decimal('1.5')
        warning = low or urgent
        result.append({
            'id': ingredient.id,
            'name': ingredient.name,
            'unit': ingredient.unit,
            'stock_quantity': float(stock),
            'reorder_level': float(reorder),
            'average_daily_usage': float(average.quantize(Decimal('0.001'))),
            'estimated_days_left': float(days_left.quantize(Decimal('0.1'))) if days_left is not None else None,
            'warning': warning,
            'status': 'urgent' if urgent else ('low' if low else 'ok'),
        })
    return result


@api_view(['GET'])
@admin_token_required
def admin_smart_finance_overview(request):
    from datetime import timedelta
    from .models import RestaurantFinancialEntry, RecurringExpenseRule

    try:
        date_from, date_to = _smart_finance_date_range(request)
    except ValueError as exc:
        return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

    days = (date_to - date_from).days + 1
    approved_statuses = [
        RestaurantFinancialEntry.STATUS_APPROVED,
        RestaurantFinancialEntry.STATUS_REIMBURSED,
    ]
    orders = Order.objects.filter(
        created_at__date__gte=date_from,
        created_at__date__lte=date_to,
    ).exclude(status=Order.STATUS_CANCELLED)
    entries = RestaurantFinancialEntry.objects.filter(
        entry_date__gte=date_from,
        entry_date__lte=date_to,
        status__in=approved_statuses,
        entry_type=RestaurantFinancialEntry.TYPE_EXPENSE,
    )

    revenue = orders.aggregate(value=Sum('total')).get('value') or Decimal('0.00')
    discounts = orders.aggregate(value=Sum('discount')).get('value') or Decimal('0.00')
    delivery_fees = orders.aggregate(value=Sum('delivery_fee')).get('value') or Decimal('0.00')
    actual_expenses = entries.aggregate(value=Sum('amount')).get('value') or Decimal('0.00')

    active_rules = RecurringExpenseRule.objects.filter(
        is_active=True,
        start_date__lte=date_to,
    ).filter(Q(end_date__isnull=True) | Q(end_date__gte=date_from)).select_related('category')
    recurring_rows = []
    recurring_allocated = Decimal('0.00')
    for rule in active_rules:
        allocation = _smart_finance_recurring_amount(rule, date_from, date_to)
        recurring_allocated += allocation
        recurring_rows.append({
            **_smart_finance_rule_payload(rule),
            'allocated_amount': float(allocation.quantize(Decimal('0.01'))),
        })

    recipe = _smart_finance_recipe_costs(date_from, date_to)
    product_cost = recipe['estimated_cost']
    product_margin = revenue - product_cost
    cash_net_before_recurring = revenue - actual_expenses
    cash_net_after_recurring = cash_net_before_recurring - recurring_allocated

    # Recurring expenses are a forecast allocation. They may duplicate cash entries
    # if the manager also registers the same bill as a normal expense.
    warnings = []
    if recurring_allocated > 0:
        warnings.append({
            'level': 'info',
            'code': 'recurring_estimate',
            'message_es': 'Los costes recurrentes se asignan como estimación. No registres el mismo recibo dos veces en Contabilidad.',
            'message_fa': 'هزینه‌های تکراری به‌صورت تخمینی تخصیص می‌یابند؛ همان قبض را دوباره در حسابداری ثبت نکنید.',
            'message_ar': 'يتم توزيع التكاليف المتكررة كتقدير؛ لا تسجل نفس الفاتورة مرتين في المحاسبة.',
        })
    if recipe['unconfigured_revenue'] > 0:
        warnings.append({
            'level': 'warning',
            'code': 'missing_recipes',
            'message_es': 'Hay ventas sin receta de coste configurada; el margen real por producto es incompleto.',
            'message_fa': 'برای بعضی فروش‌ها دستور و هزینه مواد اولیه ثبت نشده است؛ سود واقعی غذاها کامل نیست.',
            'message_ar': 'هناك مبيعات لمنتجات بدون وصفة تكلفة؛ هامش الربح الحقيقي للمنتجات غير مكتمل.',
        })

    inventory = _smart_finance_inventory_payload(date_to)
    for row in inventory:
        if row['warning']:
            warnings.append({
                'level': 'danger' if row['status'] == 'urgent' else 'warning',
                'code': 'inventory_low',
                'ingredient_id': row['id'],
                'message_es': f"{row['name']} necesita revisión de stock.",
                'message_fa': f"موجودی {row['name']} نیاز به بررسی دارد.",
                'message_ar': f"مخزون {row['name']} يحتاج إلى مراجعة.",
            })

    # Forecast: average actual cash result of the latest 14 available days.
    history_start = date_to - timedelta(days=13)
    history_orders = Order.objects.filter(
        created_at__date__gte=history_start,
        created_at__date__lte=date_to,
    ).exclude(status=Order.STATUS_CANCELLED)
    history_revenue = history_orders.aggregate(value=Sum('total')).get('value') or Decimal('0.00')
    history_expenses = RestaurantFinancialEntry.objects.filter(
        entry_date__gte=history_start,
        entry_date__lte=date_to,
        status__in=approved_statuses,
        entry_type=RestaurantFinancialEntry.TYPE_EXPENSE,
    ).aggregate(value=Sum('amount')).get('value') or Decimal('0.00')
    history_recurring = sum(
        (_smart_finance_recurring_amount(rule, history_start, date_to) for rule in active_rules),
        Decimal('0.00'),
    )
    avg_daily_net = (history_revenue - history_expenses - history_recurring) / Decimal('14')

    gross_margin_percent = (
        (product_margin / revenue * Decimal('100.00'))
        if revenue > 0 and recipe['configured_revenue'] > 0 else Decimal('0.00')
    )
    recurring_daily = recurring_allocated / Decimal(days) if days else Decimal('0.00')
    break_even_sales = (
        recurring_daily / (gross_margin_percent / Decimal('100.00'))
        if gross_margin_percent > 0 else Decimal('0.00')
    )

    expense_categories = list(
        entries.values('category__name').annotate(total=Sum('amount')).order_by('-total')
    )
    return Response({
        'period': {'date_from': date_from.isoformat(), 'date_to': date_to.isoformat(), 'days': days},
        'profit_loss': {
            'gross_sales': float(revenue),
            'discounts': float(discounts),
            'delivery_fees_collected': float(delivery_fees),
            'actual_logged_expenses': float(actual_expenses),
            'recurring_cost_allocation': float(recurring_allocated.quantize(Decimal('0.01'))),
            'cash_net_before_recurring': float(cash_net_before_recurring),
            'estimated_cash_net': float(cash_net_after_recurring),
            'estimated_recipe_cost': float(product_cost.quantize(Decimal('0.01'))),
            'estimated_product_margin': float(product_margin.quantize(Decimal('0.01'))),
            'product_margin_percent': float(gross_margin_percent.quantize(Decimal('0.01'))),
            'orders_count': orders.count(),
        },
        'forecast': {
            'average_daily_net_last_14_days': float(avg_daily_net.quantize(Decimal('0.01'))),
            'next_7_days_net': float((avg_daily_net * Decimal('7')).quantize(Decimal('0.01'))),
            'next_30_days_net': float((avg_daily_net * Decimal('30')).quantize(Decimal('0.01'))),
            'trend': 'profit' if avg_daily_net >= 0 else 'loss',
        },
        'break_even': {
            'daily_recurring_cost': float(recurring_daily.quantize(Decimal('0.01'))),
            'gross_margin_percent': float(gross_margin_percent.quantize(Decimal('0.01'))),
            'daily_sales_needed': float(break_even_sales.quantize(Decimal('0.01'))),
            'available': bool(gross_margin_percent > 0),
        },
        'recurring_costs': recurring_rows,
        'inventory': inventory,
        'expense_categories': [
            {'name': row['category__name'] or 'Sin categoría', 'amount': float(row['total'] or 0)}
            for row in expense_categories
        ],
        'warnings': warnings,
    })


@api_view(['GET', 'POST'])
@admin_token_required
def admin_smart_finance_recurring_costs(request):
    from .models import ExpenseCategory, RecurringExpenseRule
    from django.utils.dateparse import parse_date

    if request.method == 'GET':
        rows = RecurringExpenseRule.objects.select_related('category').all()
        return Response([_smart_finance_rule_payload(row) for row in rows])

    data = request.data or {}
    title = str(data.get('title') or '').strip()
    if not title:
        return Response({'detail': 'El título es obligatorio.'}, status=status.HTTP_400_BAD_REQUEST)
    try:
        amount = Decimal(str(data.get('amount') or '0'))
        if amount <= 0:
            raise ValueError('El importe debe ser mayor que cero.')
        category = None
        if data.get('category_id'):
            category = ExpenseCategory.objects.get(id=int(data['category_id']))
        start_date = parse_date(str(data.get('start_date') or '')) or timezone.localdate()
        end_date = parse_date(str(data.get('end_date') or '')) if data.get('end_date') else None
        if end_date and end_date < start_date:
            raise ValueError('La fecha final no puede ser anterior a la fecha inicial.')
        row = RecurringExpenseRule.objects.create(
            title=title,
            amount=amount,
            frequency=str(data.get('frequency') or RecurringExpenseRule.FREQUENCY_MONTHLY),
            category=category,
            paid_by=str(data.get('paid_by') or RestaurantFinancialEntry.PARTY_BBVA),
            start_date=start_date,
            end_date=end_date,
            is_active=bool(data.get('is_active', True)),
            notes=str(data.get('notes') or '').strip(),
            created_by_username=request.admin_user.get_username(),
            updated_by_username=request.admin_user.get_username(),
        )
    except (ValueError, TypeError, ExpenseCategory.DoesNotExist) as exc:
        return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
    return Response(_smart_finance_rule_payload(row), status=status.HTTP_201_CREATED)


@api_view(['PATCH', 'DELETE'])
@admin_token_required
def admin_smart_finance_recurring_cost_detail(request, rule_id):
    from .models import ExpenseCategory, RecurringExpenseRule
    from django.utils.dateparse import parse_date

    try:
        row = RecurringExpenseRule.objects.select_related('category').get(id=rule_id)
    except RecurringExpenseRule.DoesNotExist:
        return Response({'detail': 'Coste recurrente no encontrado.'}, status=status.HTTP_404_NOT_FOUND)

    if request.method == 'DELETE':
        row.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    data = request.data or {}
    try:
        if 'title' in data:
            row.title = str(data['title'] or '').strip()
        if not row.title:
            raise ValueError('El título es obligatorio.')
        if 'amount' in data:
            row.amount = Decimal(str(data['amount'] or '0'))
            if row.amount <= 0:
                raise ValueError('El importe debe ser mayor que cero.')
        if 'frequency' in data:
            row.frequency = str(data['frequency'])
        if 'category_id' in data:
            row.category = ExpenseCategory.objects.get(id=int(data['category_id'])) if data['category_id'] else None
        if 'paid_by' in data:
            row.paid_by = str(data['paid_by'])
        if 'start_date' in data:
            row.start_date = parse_date(str(data['start_date'])) or row.start_date
        if 'end_date' in data:
            row.end_date = parse_date(str(data['end_date'])) if data['end_date'] else None
        if row.end_date and row.end_date < row.start_date:
            raise ValueError('La fecha final no puede ser anterior a la fecha inicial.')
        if 'is_active' in data:
            row.is_active = bool(data['is_active'])
        if 'notes' in data:
            row.notes = str(data['notes'] or '').strip()
        row.updated_by_username = request.admin_user.get_username()
        row.full_clean()
        row.save()
    except (ValueError, TypeError, ExpenseCategory.DoesNotExist) as exc:
        return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
    return Response(_smart_finance_rule_payload(row))

# ============================================================
# v22 Real Inventory: purchases, waste, adjustments, sale usage
# ============================================================

def _inventory_decimal(value, default='0'):
    return Decimal(str(value if value not in (None, '') else default))


def _inventory_movement_payload(row):
    return {
        'id': row.id,
        'ingredient_id': row.ingredient_id,
        'ingredient_name': row.ingredient.name if hasattr(row, 'ingredient') else '',
        'unit': row.ingredient.unit if hasattr(row, 'ingredient') else '',
        'movement_type': row.movement_type,
        'quantity_delta': float(row.quantity_delta or 0),
        'unit_cost_snapshot': float(row.unit_cost_snapshot or 0),
        'total_cost': float(row.total_cost or 0),
        'iva_percent': float(getattr(row, 'iva_percent', 0) or 0),
        'iva_amount': float(getattr(row, 'iva_amount', 0) or 0),
        'total_amount_with_iva': float(getattr(row, 'total_amount_with_iva', row.total_cost) or 0),
        'supplier_name': row.supplier_name or '',
        'invoice_number': row.invoice_number or '',
        'reference': row.reference or '',
        'notes': row.notes or '',
        'occurred_at': row.occurred_at.isoformat() if row.occurred_at else '',
    }


def _inventory_apply_movement(*, ingredient_id, movement_type, quantity_delta, unit_cost_snapshot=None,
                              total_cost=None, iva_percent=None, iva_amount=None, total_amount_with_iva=None, order_item=None, financial_entry=None, supplier_name='',
                              invoice_number='', reference='', notes='', occurred_at=None, username=''):
    from .models import Ingredient, InventoryMovement

    with transaction.atomic():
        ingredient = Ingredient.objects.select_for_update().get(id=ingredient_id)
        if order_item and movement_type == InventoryMovement.TYPE_SALE:
            existing = InventoryMovement.objects.filter(
                order_item=order_item,
                ingredient=ingredient,
                movement_type=InventoryMovement.TYPE_SALE,
            ).first()
            if existing:
                return existing, False

        delta = _inventory_decimal(quantity_delta)
        unit_cost = _inventory_decimal(
            unit_cost_snapshot if unit_cost_snapshot is not None else ingredient.unit_cost,
            '0.0000'
        )
        total = _inventory_decimal(
            total_cost if total_cost is not None else abs(delta) * unit_cost,
            '0.00'
        )
        row = InventoryMovement.objects.create(
            ingredient=ingredient,
            movement_type=movement_type,
            quantity_delta=delta,
            unit_cost_snapshot=unit_cost,
            total_cost=total,
            iva_percent=_inventory_decimal(iva_percent, '0.00') if iva_percent is not None else Decimal('0.00'),
            iva_amount=_inventory_decimal(iva_amount, '0.00') if iva_amount is not None else Decimal('0.00'),
            total_amount_with_iva=_inventory_decimal(total_amount_with_iva, '0.00') if total_amount_with_iva is not None else total,
            order_item=order_item,
            financial_entry=financial_entry,
            supplier_name=str(supplier_name or '').strip(),
            invoice_number=str(invoice_number or '').strip(),
            reference=str(reference or '').strip(),
            notes=str(notes or '').strip(),
            occurred_at=occurred_at or timezone.now(),
            created_by_username=str(username or '').strip(),
        )
        ingredient.stock_quantity = _inventory_decimal(ingredient.stock_quantity) + delta
        if movement_type == InventoryMovement.TYPE_PURCHASE and unit_cost >= 0:
            ingredient.unit_cost = unit_cost
        ingredient.save(update_fields=['stock_quantity', 'unit_cost', 'updated_at'])
        return row, True


def consume_inventory_for_order(order):
    """Deduct recipe components only once, when an order reaches Delivered."""
    from .models import ProductCostProfile, InventoryMovement

    if order.status != Order.STATUS_DELIVERED:
        return {'created': 0, 'skipped': 0}

    order_items = order.items.select_related('menu_item').all()
    profiles = {
        profile.menu_item_id: profile
        for profile in ProductCostProfile.objects.prefetch_related('components__ingredient').filter(
            menu_item_id__in=[row.menu_item_id for row in order_items if row.menu_item_id]
        )
    }
    created = 0
    skipped = 0
    for order_item in order_items:
        profile = profiles.get(order_item.menu_item_id)
        if not profile:
            continue
        for component in profile.components.all():
            quantity = _inventory_decimal(component.quantity) * _inventory_decimal(order_item.quantity)
            if quantity <= 0:
                continue
            _, was_created = _inventory_apply_movement(
                ingredient_id=component.ingredient_id,
                movement_type=InventoryMovement.TYPE_SALE,
                quantity_delta=-quantity,
                unit_cost_snapshot=component.ingredient.unit_cost,
                total_cost=quantity * _inventory_decimal(component.ingredient.unit_cost),
                order_item=order_item,
                reference=f'Pedido {order.order_code}',
                notes=f'Consumo automático: {order_item.name_snapshot}',
                occurred_at=order.updated_at or timezone.now(),
                username='system',
            )
            if was_created:
                created += 1
            else:
                skipped += 1
    return {'created': created, 'skipped': skipped}


@api_view(['GET'])
@admin_token_required
def admin_inventory_real_overview(request):
    from datetime import timedelta
    from .models import Ingredient, InventoryMovement, RecipeIngredient

    today = timezone.localdate()
    days = max(1, min(90, int(request.query_params.get('days', 14) or 14)))
    start = today - timedelta(days=days - 1)

    sales = (
        Order.objects.filter(created_at__date__gte=start, created_at__date__lte=today)
        .exclude(status=Order.STATUS_CANCELLED)
        .filter(status=Order.STATUS_DELIVERED)
        .values('items__menu_item_id')
        .annotate(units=Sum('items__quantity'))
    )
    units_by_item = {row['items__menu_item_id']: _inventory_decimal(row['units']) for row in sales if row.get('items__menu_item_id')}
    daily_usage = {}
    for component in RecipeIngredient.objects.select_related('profile', 'ingredient').all():
        sold = units_by_item.get(component.profile.menu_item_id, Decimal('0.00'))
        if sold:
            daily_usage[component.ingredient_id] = daily_usage.get(component.ingredient_id, Decimal('0.00')) + (
                _inventory_decimal(component.quantity) * sold / Decimal(days)
            )

    ingredients = []
    suggestions = []
    for ingredient in Ingredient.objects.filter(is_active=True).order_by('name'):
        stock = _inventory_decimal(ingredient.stock_quantity)
        average = daily_usage.get(ingredient.id, Decimal('0.00'))
        reorder = _inventory_decimal(ingredient.reorder_level)
        days_left = (stock / average) if average > 0 else None
        target = max(reorder, average * Decimal('7'))
        suggested_quantity = max(Decimal('0.00'), target - stock)
        status_name = 'ok'
        if days_left is not None and days_left <= Decimal('1.5'):
            status_name = 'urgent'
        elif stock <= reorder:
            status_name = 'low'
        row = {
            'id': ingredient.id,
            'name': ingredient.name,
            'unit': ingredient.unit,
            'stock_quantity': float(stock),
            'unit_cost': float(ingredient.unit_cost or 0),
            'reorder_level': float(reorder),
            'supplier_name': ingredient.supplier_name or '',
            'average_daily_usage': float(average.quantize(Decimal('0.001'))),
            'estimated_days_left': float(days_left.quantize(Decimal('0.1'))) if days_left is not None else None,
            'suggested_purchase_quantity': float(suggested_quantity.quantize(Decimal('0.001'))),
            'status': status_name,
        }
        ingredients.append(row)
        if suggested_quantity > 0:
            suggestions.append(row)

    movements = InventoryMovement.objects.select_related('ingredient').order_by('-occurred_at', '-id')[:80]
    return Response({
        'days_window': days,
        'ingredients': ingredients,
        'suggested_purchase_list': suggestions,
        'recent_movements': [_inventory_movement_payload(row) for row in movements],
    })


@api_view(['POST'])
@admin_token_required
def admin_inventory_purchase(request):
    from .models import Ingredient, RestaurantFinancialEntry, ExpenseCategory, InventoryMovement
    from django.utils.dateparse import parse_datetime, parse_date

    data = request.data or {}
    try:
        ingredient = Ingredient.objects.get(id=int(data.get('ingredient_id')))
        quantity = _inventory_decimal(data.get('quantity'))
        unit_cost = _inventory_decimal(data.get('unit_cost') if data.get('unit_cost') not in (None, '') else ingredient.unit_cost, '0.0000')
        iva_percent = _inventory_decimal(data.get('iva_percent') if data.get('iva_percent') not in (None, '') else '10.00', '0.00')
        if quantity <= 0 or unit_cost < 0 or iva_percent < 0 or iva_percent > Decimal('100.00'):
            raise ValueError('Cantidad, coste unitario o IVA no válidos.')

        subtotal_amount = (quantity * unit_cost).quantize(Decimal('0.01'))
        iva_amount = (subtotal_amount * iva_percent / Decimal('100.00')).quantize(Decimal('0.01'))
        total_amount = (subtotal_amount + iva_amount).quantize(Decimal('0.01'))

        occurred = parse_datetime(str(data.get('occurred_at') or ''))
        if not occurred:
            purchase_date = parse_date(str(data.get('purchase_date') or '')) or timezone.localdate()
            occurred = timezone.make_aware(timezone.datetime.combine(purchase_date, timezone.datetime.min.time()))

        if data.get('category_id'):
            category = ExpenseCategory.objects.get(id=int(data['category_id']))
        else:
            category, _ = ExpenseCategory.objects.get_or_create(name='Materias primas')

        detail = str(data.get('notes') or '').strip()
        iva_note = f'Base: {subtotal_amount} € | IVA {iva_percent}%: {iva_amount} € | Total: {total_amount} €'
        description = f'{detail}\n{iva_note}'.strip()
        entry = RestaurantFinancialEntry.objects.create(
            entry_type=RestaurantFinancialEntry.TYPE_EXPENSE,
            title=f'Compra: {ingredient.name}',
            description=description,
            amount=total_amount,
            entry_date=occurred.date(),
            category=category,
            paid_by=str(data.get('paid_by') or RestaurantFinancialEntry.PARTY_BBVA),
            payment_method=str(data.get('payment_method') or RestaurantFinancialEntry.PAYMENT_BBVA),
            invoice_number=str(data.get('invoice_number') or '').strip(),
            status=RestaurantFinancialEntry.STATUS_APPROVED,
            created_by_username=request.admin_user.get_username(),
            updated_by_username=request.admin_user.get_username(),
        )

        movement, _ = _inventory_apply_movement(
            ingredient_id=ingredient.id,
            movement_type=InventoryMovement.TYPE_PURCHASE,
            quantity_delta=quantity,
            unit_cost_snapshot=unit_cost,
            total_cost=subtotal_amount,
            iva_percent=iva_percent,
            iva_amount=iva_amount,
            total_amount_with_iva=total_amount,
            financial_entry=entry,
            supplier_name=str(data.get('supplier_name') or ingredient.supplier_name or '').strip(),
            invoice_number=str(data.get('invoice_number') or '').strip(),
            reference='Compra de inventario',
            notes=detail,
            occurred_at=occurred,
            username=request.admin_user.get_username(),
        )
        if data.get('supplier_name'):
            ingredient.supplier_name = str(data['supplier_name']).strip()
            ingredient.save(update_fields=['supplier_name', 'updated_at'])
    except (ValueError, TypeError, Ingredient.DoesNotExist, ExpenseCategory.DoesNotExist) as exc:
        return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

    return Response({
        'success': True,
        'movement': _inventory_movement_payload(movement),
        'financial_entry_id': entry.id,
        'calculation': {
            'quantity': float(quantity),
            'unit_cost_without_iva': float(unit_cost),
            'subtotal_amount': float(subtotal_amount),
            'iva_percent': float(iva_percent),
            'iva_amount': float(iva_amount),
            'total_amount_with_iva': float(total_amount),
        },
    }, status=status.HTTP_201_CREATED)


@api_view(['POST'])
@admin_token_required
def admin_inventory_waste(request):
    from .models import Ingredient, InventoryMovement
    from django.utils.dateparse import parse_datetime

    data = request.data or {}
    try:
        ingredient = Ingredient.objects.get(id=int(data.get('ingredient_id')))
        quantity = _inventory_decimal(data.get('quantity'))
        if quantity <= 0:
            raise ValueError('La cantidad de desperdicio debe ser mayor que cero.')
        occurred = parse_datetime(str(data.get('occurred_at') or '')) or timezone.now()
        movement, _ = _inventory_apply_movement(
            ingredient_id=ingredient.id,
            movement_type=InventoryMovement.TYPE_WASTE,
            quantity_delta=-quantity,
            unit_cost_snapshot=ingredient.unit_cost,
            total_cost=quantity * _inventory_decimal(ingredient.unit_cost),
            reference='Merma / desperdicio',
            notes=str(data.get('notes') or '').strip(),
            occurred_at=occurred,
            username=request.admin_user.get_username(),
        )
    except (ValueError, TypeError, Ingredient.DoesNotExist) as exc:
        return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
    return Response({'success': True, 'movement': _inventory_movement_payload(movement)}, status=status.HTTP_201_CREATED)


@api_view(['POST'])
@admin_token_required
def admin_inventory_adjustment(request):
    from .models import Ingredient, InventoryMovement

    data = request.data or {}
    try:
        ingredient = Ingredient.objects.get(id=int(data.get('ingredient_id')))
        delta = _inventory_decimal(data.get('quantity_delta'))
        if delta == 0:
            raise ValueError('El ajuste no puede ser cero.')
        movement, _ = _inventory_apply_movement(
            ingredient_id=ingredient.id,
            movement_type=InventoryMovement.TYPE_ADJUSTMENT,
            quantity_delta=delta,
            unit_cost_snapshot=ingredient.unit_cost,
            total_cost=abs(delta) * _inventory_decimal(ingredient.unit_cost),
            reference='Ajuste manual',
            notes=str(data.get('notes') or '').strip(),
            username=request.admin_user.get_username(),
        )
    except (ValueError, TypeError, Ingredient.DoesNotExist) as exc:
        return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
    return Response({'success': True, 'movement': _inventory_movement_payload(movement)}, status=status.HTTP_201_CREATED)

# ============================================================
# v23 Profit Intelligence: product analysis, targets, partners
# ============================================================

def _profit_intelligence_range(request):
    from django.utils.dateparse import parse_date
    from datetime import timedelta
    today = timezone.localdate()
    start = parse_date(str(request.query_params.get('date_from') or '')) or (today - timedelta(days=29))
    end = parse_date(str(request.query_params.get('date_to') or '')) or today
    if start > end:
        raise ValueError('La fecha inicial no puede ser posterior a la fecha final.')
    if (end - start).days > 366:
        raise ValueError('El periodo máximo es de 366 días.')
    return start, end


def _profit_intelligence_product_rows(date_from, date_to):
    from .models import OrderItem, ProductCostProfile
    sales = (
        OrderItem.objects.filter(order__created_at__date__gte=date_from, order__created_at__date__lte=date_to)
        .exclude(order__status=Order.STATUS_CANCELLED)
        .values('menu_item_id', 'name_snapshot')
        .annotate(units=Sum('quantity'), revenue=Sum('total'))
        .order_by('-revenue')
    )
    profiles = {
        x.menu_item_id: x
        for x in ProductCostProfile.objects.select_related('menu_item').prefetch_related('components__ingredient').all()
    }
    rows = []
    for row in sales:
        profile = profiles.get(row['menu_item_id'])
        units = _inventory_decimal(row['units'])
        revenue = _inventory_decimal(row['revenue'])
        unit_cost = None
        if profile and profile.components.exists():
            ingredients_cost = sum(
                (_inventory_decimal(c.quantity) * _inventory_decimal(c.ingredient.unit_cost))
                for c in profile.components.all()
            )
            unit_cost = ingredients_cost + _inventory_decimal(profile.packaging_cost) + _inventory_decimal(profile.fixed_cost)
        total_cost = unit_cost * units if unit_cost is not None else Decimal('0.00')
        profit = revenue - total_cost if unit_cost is not None else None
        margin = (profit / revenue * Decimal('100')) if profit is not None and revenue > 0 else None
        rows.append({
            'menu_item_id': row['menu_item_id'],
            'name': row['name_snapshot'],
            'units': float(units),
            'revenue': float(revenue),
            'unit_cost': float(unit_cost.quantize(Decimal('0.0001'))) if unit_cost is not None else None,
            'total_cost': float(total_cost.quantize(Decimal('0.01'))) if unit_cost is not None else None,
            'profit': float(profit.quantize(Decimal('0.01'))) if profit is not None else None,
            'margin_percent': float(margin.quantize(Decimal('0.1'))) if margin is not None else None,
            'target_margin_percent': float(profile.target_margin_percent) if profile else None,
            'has_recipe': bool(profile and profile.components.exists()),
        })
    if not rows:
        return rows

    avg_units = sum(Decimal(str(r['units'])) for r in rows) / Decimal(len(rows))
    for row in rows:
        high_sales = Decimal(str(row['units'])) >= avg_units
        good_margin = row['margin_percent'] is not None and row['target_margin_percent'] is not None and Decimal(str(row['margin_percent'])) >= Decimal(str(row['target_margin_percent']))
        if row['margin_percent'] is None:
            row['quadrant'] = 'missing'
        elif high_sales and good_margin:
            row['quadrant'] = 'star'
        elif high_sales and not good_margin:
            row['quadrant'] = 'risk'
        elif not high_sales and good_margin:
            row['quadrant'] = 'hidden'
        else:
            row['quadrant'] = 'weak'
    return rows


@api_view(['GET'])
@admin_token_required
def admin_profit_intelligence_overview(request):
    from datetime import timedelta
    from .models import RestaurantFinancialEntry, AccountingSettings, BusinessTarget, InventoryMovement

    try:
        date_from, date_to = _profit_intelligence_range(request)
    except ValueError as exc:
        return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)

    products = _profit_intelligence_product_rows(date_from, date_to)
    revenue = sum((_inventory_decimal(x['revenue']) for x in products), Decimal('0.00'))
    estimated_product_profit = sum((_inventory_decimal(x['profit']) for x in products if x['profit'] is not None), Decimal('0.00'))
    configured_revenue = sum((_inventory_decimal(x['revenue']) for x in products if x['has_recipe']), Decimal('0.00'))

    approved = [RestaurantFinancialEntry.STATUS_APPROVED, RestaurantFinancialEntry.STATUS_REIMBURSED]
    expenses = RestaurantFinancialEntry.objects.filter(
        entry_type=RestaurantFinancialEntry.TYPE_EXPENSE,
        status__in=approved,
        entry_date__gte=date_from, entry_date__lte=date_to
    )
    total_expenses = expenses.aggregate(x=Sum('amount'))['x'] or Decimal('0.00')
    partner_rows = []
    for party in [RestaurantFinancialEntry.PARTY_SAEID, RestaurantFinancialEntry.PARTY_AHMED, RestaurantFinancialEntry.PARTY_BBVA]:
        paid = expenses.filter(paid_by=party).aggregate(x=Sum('amount'))['x'] or Decimal('0.00')
        partner_rows.append({'party': party, 'expenses_paid': float(paid)})

    waste = InventoryMovement.objects.filter(
        movement_type=InventoryMovement.TYPE_WASTE,
        occurred_at__date__gte=date_from, occurred_at__date__lte=date_to
    ).aggregate(x=Sum('total_cost'))['x'] or Decimal('0.00')
    target = BusinessTarget.current()
    settings = AccountingSettings.current()
    days = (date_to - date_from).days + 1
    month_days = 30
    revenue_target_for_period = _inventory_decimal(target.monthly_revenue_target) / Decimal(month_days) * Decimal(days)
    profit_target_for_period = _inventory_decimal(target.monthly_profit_target) / Decimal(month_days) * Decimal(days)
    net_estimate = estimated_product_profit - total_expenses
    daily_revenue_needed = max(Decimal('0.00'), (_inventory_decimal(target.monthly_revenue_target) - revenue) / Decimal(max(1, month_days - min(days, month_days))))
    daily_profit_needed = max(Decimal('0.00'), (_inventory_decimal(target.monthly_profit_target) - net_estimate) / Decimal(max(1, month_days - min(days, month_days))))

    warnings = []
    if configured_revenue < revenue:
        warnings.append({'level':'warning','es':'Hay productos vendidos sin receta de coste completa; parte del beneficio es estimada.','fa':'برخی محصولات فروخته‌شده دستور و هزینه کامل ندارند؛ بخشی از سود تخمینی است.','ar':'هناك منتجات مباعة بلا وصفة تكلفة مكتملة؛ جزء من الربح تقديري.'})
    if waste > revenue * Decimal('0.05') and revenue > 0:
        warnings.append({'level':'danger','es':'El coste de mermas supera el 5 % de las ventas del periodo.','fa':'هزینه ضایعات بیش از ۵٪ فروش این بازه است.','ar':'تكلفة الهدر تتجاوز 5٪ من مبيعات الفترة.'})
    for row in products:
        if row['quadrant'] == 'risk':
            warnings.append({'level':'warning','es':f"{row['name']} vende bien pero su margen está bajo el objetivo.",'fa':f"{row['name']} پرفروش است اما حاشیه سود آن زیر هدف است.",'ar':f"{row['name']} يبيع جيداً لكن هامش ربحه أقل من الهدف."})
            break

    return Response({
        'period': {'date_from':date_from.isoformat(),'date_to':date_to.isoformat(),'days':days},
        'summary': {
            'revenue':float(revenue), 'estimated_product_profit':float(estimated_product_profit.quantize(Decimal('0.01'))),
            'actual_expenses':float(total_expenses), 'waste_cost':float(waste),
            'estimated_net_profit':float(net_estimate.quantize(Decimal('0.01')),
            ), 'configured_revenue':float(configured_revenue),
        },
        'targets': {
            'monthly_revenue_target':float(target.monthly_revenue_target),
            'monthly_profit_target':float(target.monthly_profit_target),
            'period_revenue_target':float(revenue_target_for_period.quantize(Decimal('0.01'))),
            'period_profit_target':float(profit_target_for_period.quantize(Decimal('0.01'))),
            'daily_revenue_needed':float(daily_revenue_needed.quantize(Decimal('0.01'))),
            'daily_profit_needed':float(daily_profit_needed.quantize(Decimal('0.01'))),
        },
        'partners': partner_rows,
        'ownership': {'saeid_percent':float(settings.saeid_share_percent),'ahmed_percent':float(settings.ahmed_share_percent)},
        'products': products,
        'warnings': warnings,
    })


@api_view(['GET', 'PATCH'])
@admin_token_required
def admin_profit_intelligence_targets(request):
    from .models import BusinessTarget
    target = BusinessTarget.current()
    if request.method == 'GET':
        return Response({'monthly_revenue_target':float(target.monthly_revenue_target),'monthly_profit_target':float(target.monthly_profit_target)})
    try:
        target.monthly_revenue_target = _inventory_decimal(request.data.get('monthly_revenue_target'))
        target.monthly_profit_target = _inventory_decimal(request.data.get('monthly_profit_target'))
        if target.monthly_revenue_target < 0 or target.monthly_profit_target < 0:
            raise ValueError('Los objetivos no pueden ser negativos.')
        target.updated_by_username = request.admin_user.get_username()
        target.save()
    except (ValueError, TypeError) as exc:
        return Response({'detail':str(exc)},status=status.HTTP_400_BAD_REQUEST)
    return Response({'monthly_revenue_target':float(target.monthly_revenue_target),'monthly_profit_target':float(target.monthly_profit_target)})

# ============================================================
# v24 Management Finance Assistant & actionable alert brief
# This assistant is deterministic: it answers only from restaurant data.
# Telegram delivery is manual from the Admin panel; no background scheduler
# is added by this patch.
# ============================================================

def _management_language(value):
    return value if value in ('es', 'fa', 'ar') else 'es'


def _management_text(es, fa, ar, language):
    return {'es': es, 'fa': fa, 'ar': ar}.get(language, es)


def _management_snapshot():
    from datetime import timedelta
    from .models import RestaurantFinancialEntry, InventoryMovement, BusinessTarget

    today = timezone.localdate()
    start_week = today - timedelta(days=6)
    approved = [RestaurantFinancialEntry.STATUS_APPROVED, RestaurantFinancialEntry.STATUS_REIMBURSED]

    today_products = _profit_intelligence_product_rows(today, today)
    week_products = _profit_intelligence_product_rows(start_week, today)
    today_revenue = sum((_inventory_decimal(x['revenue']) for x in today_products), Decimal('0.00'))
    today_product_profit = sum((_inventory_decimal(x['profit']) for x in today_products if x['profit'] is not None), Decimal('0.00'))
    today_expenses = RestaurantFinancialEntry.objects.filter(
        entry_type=RestaurantFinancialEntry.TYPE_EXPENSE,
        status__in=approved,
        entry_date=today,
    ).aggregate(x=Sum('amount'))['x'] or Decimal('0.00')
    today_waste = InventoryMovement.objects.filter(
        movement_type=InventoryMovement.TYPE_WASTE,
        occurred_at__date=today,
    ).aggregate(x=Sum('total_cost'))['x'] or Decimal('0.00')
    week_revenue = sum((_inventory_decimal(x['revenue']) for x in week_products), Decimal('0.00'))
    week_profit = sum((_inventory_decimal(x['profit']) for x in week_products if x['profit'] is not None), Decimal('0.00'))
    week_waste = InventoryMovement.objects.filter(
        movement_type=InventoryMovement.TYPE_WASTE,
        occurred_at__date__gte=start_week,
        occurred_at__date__lte=today,
    ).aggregate(x=Sum('total_cost'))['x'] or Decimal('0.00')

    inventory = _smart_finance_inventory_payload(today)
    low_stock = [x for x in inventory if x.get('status') in ['urgent', 'low']]
    target = BusinessTarget.current()
    days_in_month = 30
    daily_target = _inventory_decimal(target.monthly_revenue_target) / Decimal(days_in_month)
    risks = [x for x in week_products if x.get('quadrant') == 'risk']
    stars = [x for x in week_products if x.get('quadrant') == 'star']
    weak = [x for x in week_products if x.get('quadrant') == 'weak']

    return {
        'today': today,
        'today_revenue': today_revenue,
        'today_expenses': today_expenses,
        'today_waste': today_waste,
        'today_net': today_product_profit - today_expenses,
        'week_revenue': week_revenue,
        'week_product_profit': week_profit,
        'week_waste': week_waste,
        'low_stock': low_stock,
        'risks': risks,
        'stars': stars,
        'weak': weak,
        'daily_sales_target': daily_target,
        'monthly_revenue_target': _inventory_decimal(target.monthly_revenue_target),
        'monthly_profit_target': _inventory_decimal(target.monthly_profit_target),
    }


def _management_alerts(snapshot, language):
    alerts = []
    if snapshot['daily_sales_target'] > 0 and snapshot['today_revenue'] < snapshot['daily_sales_target']:
        difference = snapshot['daily_sales_target'] - snapshot['today_revenue']
        alerts.append({
            'level': 'warning',
            'text': _management_text(
                f'Las ventas de hoy están {difference.quantize(Decimal("0.01"))} € por debajo del objetivo diario.',
                f'فروش امروز {difference.quantize(Decimal("0.01"))} € کمتر از هدف روزانه است.',
                f'مبيعات اليوم أقل من الهدف اليومي بمقدار {difference.quantize(Decimal("0.01"))} €.',
                language,
            )
        })
    for item in snapshot['low_stock'][:3]:
        qty = item.get('suggested_purchase_quantity', 0)
        alerts.append({
            'level': 'danger' if item.get('status') == 'urgent' else 'warning',
            'text': _management_text(
                f"Stock bajo: {item['name']}. Compra sugerida: {qty} {item['unit']}.",
                f"موجودی {item['name']} کم است. خرید پیشنهادی: {qty} {item['unit']}.",
                f"مخزون {item['name']} منخفض. الشراء المقترح: {qty} {item['unit']}.",
                language,
            )
        })
    if snapshot['risks']:
        item = snapshot['risks'][0]
        alerts.append({
            'level': 'warning',
            'text': _management_text(
                f"{item['name']} vende bien pero su margen está por debajo del objetivo.",
                f"{item['name']} پرفروش است اما حاشیه سود آن پایین‌تر از هدف است.",
                f"{item['name']} يبيع جيداً لكن هامش ربحه أقل من الهدف.",
                language,
            )
        })
    if snapshot['week_revenue'] > 0 and snapshot['week_waste'] > snapshot['week_revenue'] * Decimal('0.05'):
        alerts.append({
            'level': 'danger',
            'text': _management_text(
                'El coste de mermas de la semana supera el 5 % de las ventas.',
                'هزینه ضایعات این هفته بیشتر از ۵٪ فروش است.',
                'تكلفة الهدر هذا الأسبوع تتجاوز 5٪ من المبيعات.',
                language,
            )
        })
    if not alerts:
        alerts.append({
            'level': 'success',
            'text': _management_text(
                'No hay alertas críticas en este momento.',
                'در حال حاضر هشدار مهمی وجود ندارد.',
                'لا توجد تنبيهات حرجة حالياً.',
                language,
            )
        })
    return alerts


def _management_answer(question, snapshot, language):
    query = str(question or '').strip().lower()
    euro = lambda value: f"{_inventory_decimal(value).quantize(Decimal('0.01'))} €"

    if not query:
        return _management_text(
            'Pregunta por beneficio, ventas, stock, compras, mermas, productos o socios.',
            'درباره سود، فروش، انبار، خرید، ضایعات، غذاها یا شریک‌ها سؤال بپرسید.',
            'اسأل عن الربح أو المبيعات أو المخزون أو المشتريات أو الهدر أو المنتجات أو الشركاء.',
            language,
        )

    stock_keys = ['stock', 'inventario', 'compra', 'comprar', 'موجودی', 'انبار', 'خرید', 'مخزون', 'شراء']
    profit_keys = ['beneficio', 'ganancia', 'rentable', 'sud', 'سود', 'ضرر', 'ربح', 'خسارة']
    sale_keys = ['venta', 'ventas', 'facturación', 'فروش', 'مبيعات']
    waste_keys = ['merma', 'desperdicio', 'residuo', 'ضایعات', 'هدر']
    product_keys = ['producto', 'plato', 'durum', 'kebab', 'غذا', 'محصول', 'منتج', 'طبق']
    partner_keys = ['saeid', 'ahmed', 'bbva', 'شریک', 'سعید', 'احمد', 'شريك']

    if any(x in query for x in stock_keys):
        low = snapshot['low_stock']
        if low:
            lines = [f"{x['name']}: +{x.get('suggested_purchase_quantity', 0)} {x['unit']}" for x in low[:5]]
            return _management_text(
                'Compras sugeridas: ' + '; '.join(lines) + '.',
                'خریدهای پیشنهادی: ' + '؛ '.join(lines) + '.',
                'المشتريات المقترحة: ' + '؛ '.join(lines) + '.',
                language,
            )
        return _management_text('El stock actual no tiene alertas de reposición.', 'موجودی فعلی هشدار خرید ندارد.', 'المخزون الحالي لا يحتوي على تنبيهات شراء.', language)

    if any(x in query for x in waste_keys):
        return _management_text(
            f"El coste de mermas de los últimos 7 días es {euro(snapshot['week_waste'])}.",
            f"هزینه ضایعات ۷ روز اخیر {euro(snapshot['week_waste'])} است.",
            f"تكلفة الهدر خلال آخر 7 أيام هي {euro(snapshot['week_waste'])}.",
            language,
        )

    if any(x in query for x in partner_keys):
        from .models import RestaurantFinancialEntry
        approved = [RestaurantFinancialEntry.STATUS_APPROVED, RestaurantFinancialEntry.STATUS_REIMBURSED]
        today = snapshot['today']
        start = today.replace(day=1)
        rows = RestaurantFinancialEntry.objects.filter(
            entry_type=RestaurantFinancialEntry.TYPE_EXPENSE, status__in=approved,
            entry_date__gte=start, entry_date__lte=today
        ).values('paid_by').annotate(total=Sum('amount'))
        parts = {x['paid_by']: x['total'] or Decimal('0.00') for x in rows}
        return _management_text(
            f"Gastos pagados este mes — Saeid: {euro(parts.get('saeid', 0))}; Ahmed: {euro(parts.get('ahmed', 0))}; BBVA: {euro(parts.get('bbva', 0))}.",
            f"هزینه‌های پرداخت‌شده این ماه — سعید: {euro(parts.get('saeid', 0))}؛ احمد: {euro(parts.get('ahmed', 0))}؛ BBVA: {euro(parts.get('bbva', 0))}.",
            f"المصروفات المدفوعة هذا الشهر — سعيد: {euro(parts.get('saeid', 0))}؛ أحمد: {euro(parts.get('ahmed', 0))}؛ BBVA: {euro(parts.get('bbva', 0))}.",
            language,
        )

    if any(x in query for x in product_keys):
        if snapshot['stars']:
            item = snapshot['stars'][0]
            return _management_text(
                f"Producto recomendado: {item['name']}. Tiene ventas y margen buenos; conviene destacarlo.",
                f"غذای پیشنهادی: {item['name']}. فروش و حاشیه سود خوبی دارد؛ بهتر است برجسته شود.",
                f"المنتج المقترح: {item['name']}. لديه مبيعات وهامش ربح جيدان؛ من المناسب إبرازُه.",
                language,
            )
        if snapshot['risks']:
            item = snapshot['risks'][0]
            return _management_text(
                f"Revisar {item['name']}: vende bien, pero el margen está bajo el objetivo.",
                f"{item['name']} را بررسی کنید: پرفروش است اما حاشیه سود پایین‌تر از هدف است.",
                f"راجع {item['name']}: يبيع جيداً لكن الهامش أقل من الهدف.",
                language,
            )

    if any(x in query for x in profit_keys):
        return _management_text(
            f"Resultado estimado de hoy: ingresos {euro(snapshot['today_revenue'])}, gastos registrados {euro(snapshot['today_expenses'])}, beneficio neto estimado {euro(snapshot['today_net'])}.",
            f"نتیجه تخمینی امروز: فروش {euro(snapshot['today_revenue'])}، هزینه ثبت‌شده {euro(snapshot['today_expenses'])}، سود خالص تخمینی {euro(snapshot['today_net'])}.",
            f"النتيجة التقديرية اليوم: المبيعات {euro(snapshot['today_revenue'])}، المصروفات المسجلة {euro(snapshot['today_expenses'])}، صافي الربح التقديري {euro(snapshot['today_net'])}.",
            language,
        )

    if any(x in query for x in sale_keys):
        return _management_text(
            f"Ventas de hoy: {euro(snapshot['today_revenue'])}. Objetivo diario actual: {euro(snapshot['daily_sales_target'])}.",
            f"فروش امروز: {euro(snapshot['today_revenue'])}. هدف روزانه فعلی: {euro(snapshot['daily_sales_target'])}.",
            f"مبيعات اليوم: {euro(snapshot['today_revenue'])}. الهدف اليومي الحالي: {euro(snapshot['daily_sales_target'])}.",
            language,
        )

    return _management_text(
        'Puedo responder con datos sobre beneficio, ventas, inventario, compras sugeridas, mermas, productos y socios.',
        'می‌توانم با داده‌های واقعی درباره سود، فروش، انبار، خرید پیشنهادی، ضایعات، غذاها و شریک‌ها پاسخ بدهم.',
        'يمكنني الإجابة بالبيانات الفعلية عن الربح والمبيعات والمخزون والمشتريات المقترحة والهدر والمنتجات والشركاء.',
        language,
    )


@api_view(['GET'])
@admin_token_required
def admin_management_daily_brief(request):
    language = _management_language(request.query_params.get('language', 'es'))
    snapshot = _management_snapshot()
    alerts = _management_alerts(snapshot, language)
    suggestion = _management_answer('producto', snapshot, language)
    return Response({
        'date': snapshot['today'].isoformat(),
        'summary': {
            'sales': float(snapshot['today_revenue']),
            'registered_expenses': float(snapshot['today_expenses']),
            'waste_cost': float(snapshot['today_waste']),
            'estimated_net_profit': float(snapshot['today_net']),
            'daily_sales_target': float(snapshot['daily_sales_target']),
        },
        'alerts': alerts,
        'recommendation': suggestion,
    })


@api_view(['POST'])
@admin_token_required
def admin_management_assistant(request):
    language = _management_language((request.data or {}).get('language', 'es'))
    question = str((request.data or {}).get('question') or '').strip()
    snapshot = _management_snapshot()
    return Response({
        'answer': _management_answer(question, snapshot, language),
        'alerts': _management_alerts(snapshot, language),
        'data_timestamp': timezone.now().isoformat(),
    })


@api_view(['POST'])
@admin_token_required
def admin_management_send_telegram_brief(request):
    language = _management_language((request.data or {}).get('language', 'es'))
    snapshot = _management_snapshot()
    alerts = _management_alerts(snapshot, language)
    summary = _management_text(
        f"📊 Resumen de hoy\nVentas: {snapshot['today_revenue'].quantize(Decimal('0.01'))} €\nBeneficio neto estimado: {snapshot['today_net'].quantize(Decimal('0.01'))} €",
        f"📊 خلاصه امروز\nفروش: {snapshot['today_revenue'].quantize(Decimal('0.01'))} €\nسود خالص تخمینی: {snapshot['today_net'].quantize(Decimal('0.01'))} €",
        f"📊 ملخص اليوم\nالمبيعات: {snapshot['today_revenue'].quantize(Decimal('0.01'))} €\nصافي الربح التقديري: {snapshot['today_net'].quantize(Decimal('0.01'))} €",
        language,
    )
    message = 'Casa de Kebab Turco\n' + summary + '\n\n' + '\n'.join(f"• {x['text']}" for x in alerts[:4])
    ok = send_telegram_message(message)
    return Response({'success': bool(ok), 'message': message})

# ============================================================
# v25 Professional reports: preview, CSV, XLSX and PDF
# PDF generated server-side is intentionally Spanish-only because the
# standard ReportLab fonts do not reliably shape Persian/Arabic text.
# For Persian/Arabic use the multilingual on-screen report and browser print.
# ============================================================

def _professional_report_language(value):
    return value if value in ('es', 'fa', 'ar') else 'es'


def _professional_report_range(request):
    from django.utils.dateparse import parse_date
    from datetime import timedelta

    today = timezone.localdate()
    date_from = parse_date(str(request.query_params.get('date_from') or '')) or (today - timedelta(days=29))
    date_to = parse_date(str(request.query_params.get('date_to') or '')) or today
    if date_from > date_to:
        raise ValueError('La fecha inicial no puede ser posterior a la fecha final.')
    if (date_to - date_from).days > 366:
        raise ValueError('El periodo máximo es de 366 días.')
    return date_from, date_to


def _professional_report_labels(language):
    labels = {
        'es': {
            'title': 'Casa de Kebab Turco — Informe profesional',
            'period': 'Periodo',
            'financial': 'Informe financiero',
            'inventory': 'Informe de inventario',
            'profitability': 'Informe de rentabilidad',
            'partners': 'Informe de socios',
            'sales': 'Ventas',
            'expenses': 'Gastos registrados',
            'waste': 'Mermas',
            'net': 'Resultado neto estimado',
            'ingredient': 'Ingrediente',
            'stock': 'Stock actual',
            'unit_cost': 'Coste unitario',
            'movement': 'Movimiento',
            'product': 'Producto',
            'units': 'Unidades',
            'revenue': 'Ingresos',
            'cost': 'Coste',
            'profit': 'Beneficio',
            'margin': 'Margen',
            'partner': 'Pagado por',
            'amount': 'Importe',
        },
        'fa': {
            'title': 'Casa de Kebab Turco — گزارش حرفه‌ای',
            'period': 'بازه',
            'financial': 'گزارش مالی',
            'inventory': 'گزارش انبار',
            'profitability': 'گزارش سودآوری',
            'partners': 'گزارش شریک‌ها',
            'sales': 'فروش',
            'expenses': 'هزینه‌های ثبت‌شده',
            'waste': 'ضایعات',
            'net': 'نتیجه خالص تخمینی',
            'ingredient': 'ماده اولیه',
            'stock': 'موجودی فعلی',
            'unit_cost': 'هزینه واحد',
            'movement': 'گردش',
            'product': 'محصول',
            'units': 'تعداد',
            'revenue': 'درآمد',
            'cost': 'هزینه',
            'profit': 'سود',
            'margin': 'حاشیه سود',
            'partner': 'پرداخت‌کننده',
            'amount': 'مبلغ',
        },
        'ar': {
            'title': 'Casa de Kebab Turco — تقرير مهني',
            'period': 'الفترة',
            'financial': 'التقرير المالي',
            'inventory': 'تقرير المخزون',
            'profitability': 'تقرير الربحية',
            'partners': 'تقرير الشركاء',
            'sales': 'المبيعات',
            'expenses': 'المصروفات المسجلة',
            'waste': 'الهدر',
            'net': 'صافي النتيجة التقديرية',
            'ingredient': 'المادة',
            'stock': 'المخزون الحالي',
            'unit_cost': 'تكلفة الوحدة',
            'movement': 'الحركة',
            'product': 'المنتج',
            'units': 'الوحدات',
            'revenue': 'الإيرادات',
            'cost': 'التكلفة',
            'profit': 'الربح',
            'margin': 'الهامش',
            'partner': 'المدفوع بواسطة',
            'amount': 'المبلغ',
        },
    }
    return labels[language]


def _professional_report_payload(kind, date_from, date_to, language):
    from .models import RestaurantFinancialEntry, InventoryMovement, Ingredient

    labels = _professional_report_labels(language)
    approved = [
        RestaurantFinancialEntry.STATUS_APPROVED,
        RestaurantFinancialEntry.STATUS_REIMBURSED,
    ]
    title_key = kind if kind in ('financial', 'inventory', 'profitability', 'partners') else 'financial'
    result = {
        'kind': title_key,
        'title': labels[title_key],
        'labels': labels,
        'period': {'date_from': date_from.isoformat(), 'date_to': date_to.isoformat()},
        'summary': [],
        'rows': [],
    }

    if title_key == 'financial':
        orders = Order.objects.filter(
            created_at__date__gte=date_from,
            created_at__date__lte=date_to,
        ).exclude(status=Order.STATUS_CANCELLED)
        sales = orders.aggregate(total=Sum('total'))['total'] or Decimal('0.00')
        expenses = RestaurantFinancialEntry.objects.filter(
            entry_type=RestaurantFinancialEntry.TYPE_EXPENSE,
            status__in=approved,
            entry_date__gte=date_from,
            entry_date__lte=date_to,
        )
        expenses_total = expenses.aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        waste = InventoryMovement.objects.filter(
            movement_type=InventoryMovement.TYPE_WASTE,
            occurred_at__date__gte=date_from,
            occurred_at__date__lte=date_to,
        ).aggregate(total=Sum('total_cost'))['total'] or Decimal('0.00')
        result['summary'] = [
            {'label': labels['sales'], 'value': float(sales)},
            {'label': labels['expenses'], 'value': float(expenses_total)},
            {'label': labels['waste'], 'value': float(waste)},
            {'label': labels['net'], 'value': float((sales - expenses_total - waste).quantize(Decimal('0.01')))},
        ]
        result['rows'] = [
            {
                'date': row.entry_date.isoformat(),
                'title': row.title,
                'category': row.category.name if row.category_id else '',
                'paid_by': row.paid_by,
                'amount': float(row.amount),
                'invoice_number': row.invoice_number or '',
            }
            for row in expenses.select_related('category').order_by('-entry_date', '-id')
        ]
        result['columns'] = [
            ('date', 'Fecha'), ('title', 'Concepto'), ('category', 'Categoría'),
            ('paid_by', 'Pagado por'), ('amount', 'Importe (€)'), ('invoice_number', 'Factura'),
        ]

    elif title_key == 'inventory':
        ingredients = Ingredient.objects.filter(is_active=True).order_by('name')
        movements = InventoryMovement.objects.filter(
            occurred_at__date__gte=date_from,
            occurred_at__date__lte=date_to,
        ).select_related('ingredient').order_by('-occurred_at', '-id')
        result['summary'] = [
            {'label': 'Ingredientes activos', 'value': ingredients.count()},
            {'label': 'Movimientos del periodo', 'value': movements.count()},
        ]
        result['rows'] = [
            {
                'ingredient': row.ingredient.name,
                'type': row.movement_type,
                'quantity_delta': float(row.quantity_delta),
                'unit': row.ingredient.unit,
                'unit_cost': float(row.unit_cost_snapshot),
                'total_cost': float(row.total_cost),
                'supplier': row.supplier_name or '',
                'date': row.occurred_at.strftime('%Y-%m-%d %H:%M'),
            }
            for row in movements
        ]
        result['inventory_stock'] = [
            {
                'ingredient': item.name,
                'stock': float(item.stock_quantity),
                'unit': item.unit,
                'unit_cost': float(item.unit_cost),
                'reorder_level': float(item.reorder_level),
            }
            for item in ingredients
        ]
        result['columns'] = [
            ('ingredient', 'Ingrediente'), ('type', 'Movimiento'), ('quantity_delta', 'Cambio'),
            ('unit', 'Unidad'), ('unit_cost', 'Coste/u'), ('total_cost', 'Coste total'),
            ('supplier', 'Proveedor'), ('date', 'Fecha'),
        ]

    elif title_key == 'profitability':
        rows = _profit_intelligence_product_rows(date_from, date_to)
        result['rows'] = rows
        total_revenue = sum((_inventory_decimal(row['revenue']) for row in rows), Decimal('0.00'))
        total_profit = sum((_inventory_decimal(row['profit']) for row in rows if row['profit'] is not None), Decimal('0.00'))
        result['summary'] = [
            {'label': labels['revenue'], 'value': float(total_revenue)},
            {'label': labels['profit'], 'value': float(total_profit.quantize(Decimal('0.01')))},
            {'label': 'Productos con receta completa', 'value': sum(1 for row in rows if row['has_recipe'])},
        ]
        result['columns'] = [
            ('name', 'Producto'), ('units', 'Unidades'), ('revenue', 'Ingresos (€)'),
            ('unit_cost', 'Coste/u (€)'), ('total_cost', 'Coste total (€)'),
            ('profit', 'Beneficio (€)'), ('margin_percent', 'Margen %'), ('quadrant', 'Diagnóstico'),
        ]

    else:
        entries = RestaurantFinancialEntry.objects.filter(
            entry_type=RestaurantFinancialEntry.TYPE_EXPENSE,
            status__in=approved,
            entry_date__gte=date_from,
            entry_date__lte=date_to,
        ).values('paid_by').annotate(amount=Sum('amount')).order_by('paid_by')
        result['rows'] = [{'party': row['paid_by'], 'amount': float(row['amount'] or 0)} for row in entries]
        result['summary'] = [{'label': 'Total gastos de socios y BBVA', 'value': float(sum((_inventory_decimal(x['amount']) for x in result['rows']), Decimal('0.00')))}]
        result['columns'] = [('party', 'Pagado por'), ('amount', 'Importe (€)')]

    return result


def _professional_csv_response(payload):
    import csv
    from io import StringIO
    from django.http import HttpResponse

    stream = StringIO()
    writer = csv.writer(stream)
    writer.writerow([payload['title']])
    writer.writerow([f"{payload['period']['date_from']} — {payload['period']['date_to']}"])
    writer.writerow([])
    for row in payload['summary']:
        writer.writerow([row['label'], row['value']])
    writer.writerow([])
    writer.writerow([title for _, title in payload['columns']])
    for row in payload['rows']:
        writer.writerow([row.get(key, '') for key, _ in payload['columns']])
    response = HttpResponse(stream.getvalue().encode('utf-8-sig'), content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f"attachment; filename=casadekebab_{payload['kind']}_{payload['period']['date_from']}_{payload['period']['date_to']}.csv"
    return response


def _professional_xlsx_response(payload):
    from io import BytesIO
    from django.http import HttpResponse
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        return Response({'detail': 'Falta openpyxl. Instala: pip install openpyxl'}, status=status.HTTP_503_SERVICE_UNAVAILABLE)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = payload['kind'][:31]
    sheet.append([payload['title']])
    sheet.append([f"{payload['period']['date_from']} — {payload['period']['date_to']}"])
    sheet.append([])
    for row in payload['summary']:
        sheet.append([row['label'], row['value']])
    sheet.append([])
    sheet.append([title for _, title in payload['columns']])
    for row in payload['rows']:
        sheet.append([row.get(key, '') for key, _ in payload['columns']])

    for cell in sheet[1]:
        cell.font = Font(bold=True, size=14)
    header_row = 4 + len(payload['summary']) + 1
    for cell in sheet[header_row]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='8F1D18')
        cell.alignment = Alignment(horizontal='center')
    for column in range(1, sheet.max_column + 1):
        width = min(42, max(12, max(len(str(sheet.cell(row, column).value or '')) for row in range(1, sheet.max_row + 1)) + 2))
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.freeze_panes = f'A{header_row + 1}'

    output = BytesIO()
    workbook.save(output)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f"attachment; filename=casadekebab_{payload['kind']}_{payload['period']['date_from']}_{payload['period']['date_to']}.xlsx"
    return response


def _professional_pdf_response(payload):
    from io import BytesIO
    from django.http import HttpResponse
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    except ImportError:
        return Response({'detail': 'Falta reportlab. Instala: pip install reportlab'}, status=status.HTTP_503_SERVICE_UNAVAILABLE)

    # Use Spanish PDF labels to ensure reliable embedded-font output.
    spanish_payload = _professional_report_payload(payload['kind'], timezone.datetime.fromisoformat(payload['period']['date_from']).date(), timezone.datetime.fromisoformat(payload['period']['date_to']).date(), 'es')
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4, rightMargin=14*mm, leftMargin=14*mm, topMargin=14*mm, bottomMargin=14*mm)
    styles = getSampleStyleSheet()
    story = [
        Paragraph('Casa de Kebab Turco', styles['Title']),
        Paragraph(spanish_payload['title'], styles['Heading2']),
        Paragraph(f"Periodo: {spanish_payload['period']['date_from']} — {spanish_payload['period']['date_to']}", styles['Normal']),
        Spacer(1, 8),
    ]
    summary_rows = [[row['label'], f"{_inventory_decimal(row['value']).quantize(Decimal('0.01')) if isinstance(row['value'], (int, float)) else row['value']}"] for row in spanish_payload['summary']]
    if summary_rows:
        table = Table(summary_rows, colWidths=[115*mm, 55*mm])
        table.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 0.3, colors.HexColor('#D7C7BB')),
            ('BACKGROUND', (0,0), (0,-1), colors.HexColor('#F7F0EA')),
            ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
            ('PADDING', (0,0), (-1,-1), 6),
        ]))
        story.extend([table, Spacer(1, 10)])

    raw_rows = [[title for _, title in spanish_payload['columns']]]
    for row in spanish_payload['rows'][:200]:
        raw_rows.append([str(row.get(key, ''))[:46] for key, _ in spanish_payload['columns']])
    if len(raw_rows) > 1:
        widths = [170*mm / max(1, len(spanish_payload['columns']))] * len(spanish_payload['columns'])
        table = Table(raw_rows, colWidths=widths, repeatRows=1)
        table.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 0.25, colors.HexColor('#DDDDDD')),
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#8F1D18')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 7),
            ('PADDING', (0,0), (-1,-1), 4),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ]))
        story.append(table)
    doc.build(story)
    response = HttpResponse(output.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f"attachment; filename=casadekebab_{payload['kind']}_{payload['period']['date_from']}_{payload['period']['date_to']}.pdf"
    return response


@api_view(['GET'])
@admin_token_required
def admin_professional_reports_preview(request):
    try:
        date_from, date_to = _professional_report_range(request)
    except ValueError as exc:
        return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
    kind = str(request.query_params.get('kind') or 'financial').strip()
    language = _professional_report_language(request.query_params.get('language') or 'es')
    return Response(_professional_report_payload(kind, date_from, date_to, language))


@api_view(['GET'])
@admin_token_required
def admin_professional_reports_export(request):
    try:
        date_from, date_to = _professional_report_range(request)
    except ValueError as exc:
        return Response({'detail': str(exc)}, status=status.HTTP_400_BAD_REQUEST)
    kind = str(request.query_params.get('kind') or 'financial').strip()
    language = _professional_report_language(request.query_params.get('language') or 'es')
    export_format = str(request.query_params.get('format') or 'csv').strip().lower()
    payload = _professional_report_payload(kind, date_from, date_to, language)
    if export_format == 'csv':
        return _professional_csv_response(payload)
    if export_format == 'xlsx':
        return _professional_xlsx_response(payload)
    if export_format == 'pdf':
        return _professional_pdf_response(payload)
    return Response({'detail': 'Formato no válido. Usa csv, xlsx o pdf.'}, status=status.HTTP_400_BAD_REQUEST)

